from __future__ import annotations

import csv
import hashlib
import io
import re
import uuid
from datetime import date
from typing import Any

from argon2 import PasswordHasher
from openpyxl import load_workbook

from campus_service import CampusService, PermissionDenied, ValidationError
from database import LearningDatabase
from config import get_student_default_password


class TeacherService:
    def __init__(self, db: LearningDatabase, campus: CampusService):
        self.db = db
        self.campus = campus
        self.passwords = PasswordHasher()

    @staticmethod
    def require_teacher(actor: dict[str, Any]) -> str:
        if actor.get("role") != "teacher" or actor.get("status") != "active":
            raise PermissionDenied("仅在职教师可以执行该操作")
        return str(actor["user_id"])

    def list_courses(self, actor: dict[str, Any]) -> list[dict[str, Any]]:
        teacher_id = self.require_teacher(actor)
        return self.db.fetch_all(
            """SELECT c.*,
                      (SELECT COUNT(*) FROM classes cl WHERE cl.course_id=c.course_id) class_count,
                      (SELECT COUNT(*) FROM course_documents d WHERE d.course_id=c.course_id) document_count
               FROM courses c WHERE c.course_type='shared_course' AND c.owner_id=?
               ORDER BY c.updated_at DESC""",
            (teacher_id,),
        )

    def create_course(self, actor: dict[str, Any], name: str, description: str = "") -> dict[str, Any]:
        teacher_id = self.require_teacher(actor)
        return self.campus.create_course(name, "shared_course", teacher_id, "teacher", description, "enrolled")

    def list_terms(self, actor: dict[str, Any]) -> list[dict[str, Any]]:
        teacher_id = self.require_teacher(actor)
        return self.db.fetch_all("SELECT * FROM terms WHERE owner_id=? ORDER BY created_at DESC", (teacher_id,))

    def create_term(self, actor: dict[str, Any], name: str, starts_on: date | None = None,
                    ends_on: date | None = None, academic_year: str = "",
                    teaching_period: str = "") -> dict[str, Any]:
        teacher_id = self.require_teacher(actor)
        name = name.strip()
        if not name:
            raise ValidationError("学期名称不能为空")
        if starts_on and ends_on and starts_on > ends_on:
            raise ValidationError("学期结束日期不能早于开始日期")
        academic_year = academic_year.strip()[:32]
        teaching_period = teaching_period.strip()[:64] or name
        if not academic_year:
            match = re.search(r"(20\d{2})(?:\s*[-—至/]\s*(20\d{2}))?", name)
            if match:
                academic_year = (
                    f"{match.group(1)}-{match.group(2)}" if match.group(2) else match.group(1)
                )
        term_id = f"term_{uuid.uuid4().hex[:12]}"
        try:
            self.db.execute(
                """INSERT INTO terms(
                       term_id,term_name,owner_id,starts_on,ends_on,academic_year,teaching_period
                   ) VALUES(?,?,?,?,?,?,?)""",
                (term_id, name, teacher_id, starts_on.isoformat() if starts_on else None,
                 ends_on.isoformat() if ends_on else None, academic_year, teaching_period),
            )
        except Exception as exc:
            if "UNIQUE" in str(exc).upper():
                raise ValidationError("该学期名称已存在") from exc
            raise
        return self.db.fetch_one("SELECT * FROM terms WHERE term_id=?", (term_id,)) or {}

    def list_classes(self, actor: dict[str, Any], course_id: str | None = None) -> list[dict[str, Any]]:
        teacher_id = self.require_teacher(actor)
        params: tuple[Any, ...] = (teacher_id,)
        condition = "cl.teacher_id=?"
        if course_id:
            condition += " AND cl.course_id=?"
            params += (course_id,)
        return self.db.fetch_all(
            f"""SELECT cl.*,c.course_name,t.term_name,
                       (SELECT COUNT(*) FROM class_memberships m WHERE m.class_id=cl.class_id AND m.status='active') member_count
                FROM classes cl JOIN courses c ON c.course_id=cl.course_id
                JOIN terms t ON t.term_id=cl.term_id WHERE {condition}
                ORDER BY COALESCE(NULLIF(t.academic_year,''),t.term_name) DESC,
                         t.teaching_period,cl.class_variant,cl.teaching_time_slot,cl.class_name""",
            params,
        )

    def create_class(self, actor: dict[str, Any], course_id: str, term_id: str,
                     class_name: str, class_variant: str = "",
                     teaching_time_slot: str = "", campus: str = "",
                     cohort_year: str = "", major: str = "",
                     teaching_level: str = "") -> dict[str, Any]:
        teacher_id = self.require_teacher(actor)
        course = self.campus.require_access(course_id, teacher_id, "teacher")
        if course["course_type"] != "shared_course" or course["owner_id"] != teacher_id:
            raise PermissionDenied("只能为自己的共享课程创建教学班")
        term = self.db.fetch_one("SELECT * FROM terms WHERE term_id=? AND owner_id=?", (term_id, teacher_id))
        if not term:
            raise PermissionDenied("无权使用该学期")
        class_name = class_name.strip()
        if not class_name:
            raise ValidationError("教学班名称不能为空")
        class_variant = class_variant.strip()[:100]
        teaching_time_slot = teaching_time_slot.strip()[:120]
        campus = campus.strip()[:100]
        cohort_year = cohort_year.strip()[:32]
        major = major.strip()[:120]
        teaching_level = teaching_level.strip()[:100]
        class_id = f"class_{uuid.uuid4().hex[:12]}"
        self.db.execute(
            """INSERT INTO classes(
                   class_id,course_id,term_id,class_name,teacher_id,class_variant,teaching_time_slot,
                   campus,cohort_year,major,teaching_level
               ) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
            (class_id, course_id, term_id, class_name, teacher_id,
             class_variant, teaching_time_slot, campus, cohort_year, major, teaching_level),
        )
        rows = self.list_classes(actor, course_id)
        return next(row for row in rows if row["class_id"] == class_id)

    def require_class(self, actor: dict[str, Any], class_id: str) -> dict[str, Any]:
        teacher_id = self.require_teacher(actor)
        row = self.db.fetch_one("SELECT * FROM classes WHERE class_id=? AND teacher_id=?", (class_id, teacher_id))
        if not row:
            raise PermissionDenied("无权访问该教学班")
        return row

    def list_members(self, actor: dict[str, Any], class_id: str) -> list[dict[str, Any]]:
        self.require_class(actor, class_id)
        return self.db.fetch_all(
            """SELECT m.student_id AS user_id,u.student_number,u.username,u.display_name,
                      u.must_change_password,u.password_changed_at,
                      m.anonymous_id,m.status,m.created_at
               FROM class_memberships m LEFT JOIN users u ON u.user_id=m.student_id
               WHERE m.class_id=? ORDER BY m.created_at""",
            (class_id,),
        )

    def reset_student_password(self, actor: dict[str, Any], class_id: str,
                               student_id: str, new_password: str) -> dict[str, Any]:
        self.require_class(actor, class_id)
        if len(new_password) < 10:
            raise ValidationError("新密码至少需要 10 个字符")
        member = self.db.fetch_one(
            """SELECT u.user_id,u.role,u.username,u.student_number,u.display_name
                 FROM class_memberships m JOIN users u ON u.user_id=m.student_id
                WHERE m.class_id=? AND m.student_id=? AND m.status='active'""",
            (class_id, student_id),
        )
        if not member or member["role"] != "student":
            raise PermissionDenied("只能重置当前教学班在册学生的密码")
        with self.db.connect() as conn:
            conn.execute(
                """UPDATE users SET password_hash=?,must_change_password=1,
                       password_changed_at=NULL,updated_at=CURRENT_TIMESTAMP WHERE user_id=?""",
                (self.passwords.hash(new_password), student_id),
            )
            conn.execute("DELETE FROM refresh_tokens WHERE user_id=?", (student_id,))
        return {
            "user_id": student_id,
            "student_number": member.get("student_number") or member.get("username"),
            "display_name": member.get("display_name") or "",
            "must_change_password": 1,
        }

    def add_members(self, actor: dict[str, Any], class_id: str, student_ids: list[str]) -> dict[str, Any]:
        return self.import_members(actor, class_id, [
            {"student_number": value, "display_name": ""} for value in student_ids
        ])

    @staticmethod
    def _valid_student_number(value: str) -> bool:
        return bool(re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9_.-]{2,63}", value))

    def import_members(self, actor: dict[str, Any], class_id: str,
                       rows: list[dict[str, Any]]) -> dict[str, Any]:
        scope = self.require_class(actor, class_id)
        if not rows:
            raise ValidationError("学生名单不能为空")
        if len(rows) > 5000:
            raise ValidationError("单次最多导入 5000 行")
        default_password = get_student_default_password()
        if not default_password:
            raise ValidationError("未配置 ZHIJIAO_STUDENT_DEFAULT_PASSWORD，无法创建学生账号")

        results: list[dict[str, Any]] = []
        for index, raw in enumerate(rows, 1):
            number = str(raw.get("student_number") or "").strip().lower()
            name = str(raw.get("display_name") or "").strip()[:100]
            base = {"row": index, "student_number": number, "display_name": name}
            if not self._valid_student_number(number):
                results.append({**base, "status": "invalid", "message": "学号格式无效"})
                continue
            try:
                with self.db.connect() as conn:
                    found = conn.execute(
                        "SELECT * FROM users WHERE student_number=? OR username=? ORDER BY student_number IS NOT NULL DESC LIMIT 1",
                        (number, number),
                    ).fetchone()
                    created = False
                    if found:
                        user = dict(found)
                        if user["role"] != "student":
                            results.append({**base, "status": "conflict", "message": "该学号已被非学生账号占用"})
                            continue
                        user_id = user["user_id"]
                        if not user.get("student_number"):
                            conn.execute("UPDATE users SET student_number=? WHERE user_id=?", (number, user_id))
                        if name and not user.get("display_name"):
                            conn.execute("UPDATE users SET display_name=?,updated_at=CURRENT_TIMESTAMP WHERE user_id=?", (name, user_id))
                    else:
                        user_id = f"s_{uuid.uuid4().hex[:16]}"
                        conn.execute(
                            """INSERT INTO users(user_id,username,password_hash,role,display_name,student_number,must_change_password)
                               VALUES(?,?,?,'student',?,?,1)""",
                            (user_id, number, self.passwords.hash(default_password), name, number),
                        )
                        created = True
                    # Convert a legacy membership/enrollment that stored the school
                    # number directly into the canonical internal user id.
                    if user_id != number:
                        legacy_member = conn.execute(
                            "SELECT status FROM class_memberships WHERE class_id=? AND student_id=?",
                            (class_id, number),
                        ).fetchone()
                        if legacy_member:
                            conn.execute("DELETE FROM class_memberships WHERE class_id=? AND student_id=?", (class_id, number))
                        conn.execute(
                            "DELETE FROM course_enrollments WHERE course_id=? AND student_id=?",
                            (scope["course_id"], number),
                        )
                    existing = conn.execute(
                        "SELECT status FROM class_memberships WHERE class_id=? AND student_id=?",
                        (class_id, user_id),
                    ).fetchone()
                    if existing and existing["status"] == "active":
                        status = "already_member"
                    else:
                        anon = hashlib.sha256(f"{class_id}:{user_id}".encode()).hexdigest()[:16]
                        conn.execute(
                            """INSERT INTO class_memberships(class_id,student_id,anonymous_id,status) VALUES(?,?,?,'active')
                               ON CONFLICT(class_id,student_id) DO UPDATE SET status='active'""",
                            (class_id, user_id, anon),
                        )
                        conn.execute(
                            "INSERT OR IGNORE INTO course_enrollments(course_id,student_id) VALUES(?,?)",
                            (scope["course_id"], user_id),
                        )
                        status = "created" if created else "reused"
                results.append({**base, "user_id": user_id, "status": status, "message": ""})
            except Exception as exc:
                results.append({**base, "status": "conflict", "message": str(exc)[:160]})
        summary = {key: sum(1 for item in results if item["status"] == key)
                   for key in ("created", "reused", "already_member", "conflict", "invalid")}
        return {
            "total": len(results), "imported": summary["created"] + summary["reused"],
            "summary": summary, "results": results, "members": self.list_members(actor, class_id),
        }

    def import_member_file(self, actor: dict[str, Any], class_id: str, filename: str,
                           data: bytes) -> dict[str, Any]:
        suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""
        if suffix == "csv":
            try:
                parsed = list(csv.DictReader(io.StringIO(data.decode("utf-8-sig"))))
            except UnicodeDecodeError as exc:
                raise ValidationError("CSV 文件必须使用 UTF-8 编码") from exc
        elif suffix == "xlsx":
            try:
                values = load_workbook(io.BytesIO(data), read_only=True, data_only=True).active.iter_rows(values_only=True)
                headers = [str(value or "").strip() for value in next(values)]
                parsed = [dict(zip(headers, row)) for row in values]
            except (StopIteration, ValueError) as exc:
                raise ValidationError("XLSX 文件为空或格式无效") from exc
        else:
            raise ValidationError("只支持 CSV 或 XLSX 名单文件")
        if len(parsed) > 5000:
            raise ValidationError("单次最多导入 5000 行")
        aliases = {"学号": "student_number", "student_number": "student_number",
                   "姓名": "display_name", "display_name": "display_name"}
        normalized = [{aliases.get(str(key).strip(), str(key).strip()): value
                       for key, value in row.items()} for row in parsed]
        if not normalized or "student_number" not in normalized[0]:
            raise ValidationError("名单缺少“学号”或 student_number 列")
        return self.import_members(actor, class_id, normalized)
