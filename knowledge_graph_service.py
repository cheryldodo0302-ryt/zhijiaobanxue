from __future__ import annotations

import hashlib
import json
import re
import unicodedata
import uuid
from pathlib import Path, PurePosixPath
from typing import Any, BinaryIO

from openpyxl import load_workbook

from campus_service import MAX_UPLOAD_BYTES, CampusService, NotFound, PermissionDenied, ValidationError
from database import LearningDatabase


GRAPH_FILE_MIMES = {
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/octet-stream",
    },
    ".xls": {"application/vnd.ms-excel", "application/octet-stream"},
}
FILE_KINDS = {"auto", "nodes", "relations", "definitions", "ignore"}
NODE_STATUSES = {"draft", "approved", "rejected"}
RELATION_KINDS = {"part_of", "prerequisite", "progression", "parallel", "related"}
RELATION_MAP = {
    "整部": ("part_of", "整体—部分"),
    "整体": ("part_of", "整体—部分"),
    "依赖": ("prerequisite", "前置关系"),
    "递进": ("progression", "后续进阶"),
    "共生": ("parallel", "双向并列"),
}


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def _loads(value: str | None, default: Any) -> Any:
    try:
        return json.loads(value or "")
    except (TypeError, json.JSONDecodeError):
        return default


def _clean(value: Any, limit: int = 500) -> str:
    return " ".join(str(value or "").replace("\x00", "").split())[:limit]


def _normalized_title(value: Any) -> str:
    text = unicodedata.normalize("NFKC", _clean(value, 300)).casefold()
    return re.sub(r"[\s·•,，。；;：:（）()《》\[\]【】_—-]+", "", text)


def _safe_relative_path(value: str, fallback: str) -> str:
    normalized = str(value or fallback).replace("\\", "/").strip(" /")
    path = PurePosixPath(normalized)
    if path.is_absolute() or not path.parts or any(part in {"", ".", ".."} for part in path.parts):
        raise ValidationError("文件夹路径不安全")
    return "/".join(_clean(part, 120) for part in path.parts)


class KnowledgeGraphService:
    """Independent, versioned course graph with reviewed imports and immutable publication snapshots."""

    def __init__(self, db: LearningDatabase, campus: CampusService):
        self.db = db
        self.campus = campus
        self.storage_root = (campus.storage_dir / "knowledge_graph").resolve()
        self.storage_root.mkdir(parents=True, exist_ok=True)

    def _teacher_course(self, actor: dict[str, Any], course_id: str) -> dict[str, Any]:
        if actor.get("role") != "teacher":
            raise PermissionDenied("仅教师可以管理知识图谱")
        course = self.campus.require_access(course_id, str(actor["user_id"]), "teacher")
        if course["course_type"] != "shared_course" or course["owner_id"] != actor["user_id"]:
            raise PermissionDenied("只能管理自己的共享课程知识图谱")
        return course

    def _graph(self, actor: dict[str, Any], course_id: str, *, create: bool = True) -> dict[str, Any]:
        self._teacher_course(actor, course_id)
        row = self.db.fetch_one("SELECT * FROM knowledge_graphs WHERE course_id=?", (course_id,))
        if not row and create:
            graph_id = f"kg_{uuid.uuid4().hex}"
            self.db.execute(
                "INSERT INTO knowledge_graphs(graph_id,course_id,created_by) VALUES(?,?,?)",
                (graph_id, course_id, actor["user_id"]),
            )
            row = self.db.fetch_one("SELECT * FROM knowledge_graphs WHERE graph_id=?", (graph_id,))
        if not row:
            raise NotFound("课程尚未创建知识图谱")
        return row

    def _require_batch(self, actor: dict[str, Any], batch_id: str) -> dict[str, Any]:
        row = self.db.fetch_one(
            """SELECT b.* FROM knowledge_graph_import_batches b JOIN courses c USING(course_id)
               WHERE b.batch_id=? AND c.owner_id=?""", (batch_id, actor["user_id"]),
        )
        if not row or actor.get("role") != "teacher":
            raise PermissionDenied("无权访问该知识图谱导入批次")
        return row

    def create_import_batch(self, actor: dict[str, Any], course_id: str,
                            batch_name: str = "") -> dict[str, Any]:
        graph = self._graph(actor, course_id)
        batch_id = f"kgib_{uuid.uuid4().hex}"
        self.db.execute(
            """INSERT INTO knowledge_graph_import_batches(
                   batch_id,graph_id,course_id,batch_name,created_by
               ) VALUES(?,?,?,?,?)""",
            (batch_id, graph["graph_id"], course_id,
             _clean(batch_name, 160) or "知识图谱整包导入", actor["user_id"]),
        )
        return self.get_import_batch(actor, batch_id)

    def get_import_batch(self, actor: dict[str, Any], batch_id: str) -> dict[str, Any]:
        batch = self._require_batch(actor, batch_id)
        files = self.db.fetch_all(
            "SELECT * FROM knowledge_graph_import_files WHERE batch_id=? ORDER BY relative_path,file_id",
            (batch_id,),
        )
        for row in files:
            row.pop("stored_path", None)
            row.pop("sha256", None)
            row["risk_codes"] = _loads(row.pop("risk_codes_json", "[]"), [])
        return {**batch, "files": files}

    def add_import_file(self, actor: dict[str, Any], batch_id: str, name: str,
                        mime_type: str, stream: BinaryIO, *, relative_path: str = "") -> dict[str, Any]:
        batch = self._require_batch(actor, batch_id)
        if batch["status"] != "staging":
            raise ValidationError("该导入批次已经提交")
        original_name = Path(str(name or "graph.xlsx")).name
        suffix = Path(original_name).suffix.lower()
        if suffix not in GRAPH_FILE_MIMES or mime_type not in GRAPH_FILE_MIMES[suffix]:
            raise ValidationError("知识图谱整包只接受 XLS 或 XLSX 文件")
        relative = _safe_relative_path(relative_path, original_name)
        file_id = f"kgif_{uuid.uuid4().hex}"
        root = (self.storage_root / batch_id).resolve()
        root.mkdir(parents=True, exist_ok=True)
        stored = (root / f"{file_id}{suffix}").resolve()
        if root not in stored.parents:
            raise ValidationError("知识图谱文件存储路径无效")
        digest = hashlib.sha256()
        size = 0
        with stored.open("wb") as target:
            while True:
                chunk = stream.read(1024 * 1024)
                if not chunk:
                    break
                size += len(chunk)
                if size > MAX_UPLOAD_BYTES:
                    stored.unlink(missing_ok=True)
                    raise ValidationError(f"文件不能超过 {MAX_UPLOAD_BYTES // 1024 // 1024}MB")
                digest.update(chunk)
                target.write(chunk)
        if not size:
            stored.unlink(missing_ok=True)
            raise ValidationError("不能上传空文件")
        sha256 = digest.hexdigest()
        if self.db.fetch_one(
            "SELECT 1 ok FROM knowledge_graph_import_files WHERE batch_id=? AND sha256=?",
            (batch_id, sha256),
        ):
            stored.unlink(missing_ok=True)
            raise ValidationError("该批次中已存在内容相同的文件")
        suggested = self._suggest_kind(original_name)
        self.db.execute(
            """INSERT INTO knowledge_graph_import_files(
                   file_id,batch_id,original_name,relative_path,stored_path,mime_type,size_bytes,
                   sha256,suggested_kind,confirmed_kind
               ) VALUES(?,?,?,?,?,?,?,?,?,?)""",
            (file_id, batch_id, original_name, relative, str(stored), mime_type, size,
             sha256, suggested, suggested),
        )
        self.db.execute(
            """UPDATE knowledge_graph_import_batches SET file_count=(
                   SELECT COUNT(*) FROM knowledge_graph_import_files WHERE batch_id=?
               ),updated_at=CURRENT_TIMESTAMP WHERE batch_id=?""", (batch_id, batch_id),
        )
        return next(row for row in self.get_import_batch(actor, batch_id)["files"] if row["file_id"] == file_id)

    @staticmethod
    def _suggest_kind(name: str) -> str:
        lowered = name.casefold()
        if "关系" in lowered:
            return "relations"
        if "清单" in lowered or "知识点" in lowered:
            return "nodes"
        if "定义" in lowered:
            return "definitions"
        return "auto"

    def update_import_file(self, actor: dict[str, Any], batch_id: str, file_id: str,
                           kind: str) -> dict[str, Any]:
        batch = self._require_batch(actor, batch_id)
        if batch["status"] != "staging" or kind not in FILE_KINDS:
            raise ValidationError("文件分类无效或批次已提交")
        self.db.execute(
            "UPDATE knowledge_graph_import_files SET confirmed_kind=?,updated_at=CURRENT_TIMESTAMP WHERE file_id=? AND batch_id=?",
            (kind, file_id, batch_id),
        )
        return next(row for row in self.get_import_batch(actor, batch_id)["files"] if row["file_id"] == file_id)

    @staticmethod
    def _sheet_rows(path: Path) -> list[tuple[str, list[list[str]]]]:
        if path.suffix.lower() == ".xlsx":
            workbook = load_workbook(path, read_only=True, data_only=True)
            result = [(sheet.title, [[_clean(value, 2000) for value in row]
                                      for row in sheet.iter_rows(values_only=True)])
                      for sheet in workbook.worksheets]
            workbook.close()
            return result
        import xlrd
        workbook = xlrd.open_workbook(str(path), on_demand=True)
        result = [(sheet.name, [[_clean(sheet.cell_value(row, col), 2000)
                                 for col in range(sheet.ncols)] for row in range(sheet.nrows)])
                  for sheet in workbook.sheets()]
        workbook.release_resources()
        return result

    @staticmethod
    def _header_index(rows: list[list[str]], required: set[str]) -> int | None:
        for index, row in enumerate(rows[:30]):
            if required.issubset(set(row)):
                return index
        return None

    def _parse_file(self, row: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], dict[str, str]]:
        path = Path(row["stored_path"]).resolve()
        batch_root = (self.storage_root / row["batch_id"]).resolve()
        if batch_root not in path.parents or not path.is_file():
            raise ValidationError("导入文件位置无效")
        nodes: list[dict[str, Any]] = []
        relations: list[dict[str, Any]] = []
        definitions: dict[str, str] = {}
        confirmed = row["confirmed_kind"]
        for sheet_name, rows in self._sheet_rows(path):
            node_header = self._header_index(rows, {"知识点名称"})
            relation_header = self._header_index(rows, {"知识点1", "知识关系", "知识点2"})
            definition_header = self._header_index(rows, {"关系名称", "含义"})
            if node_header is not None and confirmed in {"auto", "nodes"}:
                header = rows[node_header]
                columns = {value: index for index, value in enumerate(header) if value}
                for row_number, values in enumerate(rows[node_header + 1:], node_header + 2):
                    title = values[columns["知识点名称"]] if columns["知识点名称"] < len(values) else ""
                    if not title:
                        continue
                    get = lambda key: values[columns[key]] if key in columns and columns[key] < len(values) else ""
                    nodes.append({
                        "title": title, "is_key": bool(get("是否重点")),
                        "is_difficult": bool(get("是否难点")), "is_exam": bool(get("是否考点")),
                        "notes": get("备注"),
                        "source": {"file_id": row["file_id"], "file": row["original_name"],
                                   "sheet": sheet_name, "row": row_number},
                    })
            if relation_header is not None and confirmed in {"auto", "relations"}:
                header = rows[relation_header]
                columns = {value: index for index, value in enumerate(header) if value}
                for row_number, values in enumerate(rows[relation_header + 1:], relation_header + 2):
                    get = lambda key: values[columns[key]] if columns[key] < len(values) else ""
                    source, relation, target = get("知识点1"), get("知识关系"), get("知识点2")
                    if source and relation and target:
                        relations.append({"source": source, "target": target, "raw_kind": relation,
                                          "source_ref": {"file_id": row["file_id"], "file": row["original_name"],
                                                         "sheet": sheet_name, "row": row_number}})
            if definition_header is not None and confirmed in {"auto", "relations", "definitions"}:
                header = rows[definition_header]
                columns = {value: index for index, value in enumerate(header) if value}
                for values in rows[definition_header + 1:]:
                    name = values[columns["关系名称"]] if columns["关系名称"] < len(values) else ""
                    meaning = values[columns["含义"]] if columns["含义"] < len(values) else ""
                    if name:
                        definitions[name] = meaning
        return nodes, relations, definitions

    def _upsert_node(self, graph_id: str, item: dict[str, Any], *, origin: str = "file") -> dict[str, Any]:
        normalized = _normalized_title(item["title"])
        if not normalized:
            raise ValidationError("知识点名称不能为空")
        existing = self.db.fetch_one(
            "SELECT * FROM knowledge_graph_nodes WHERE graph_id=? AND normalized_title=?",
            (graph_id, normalized),
        )
        if existing:
            risks = _loads(existing["risk_codes_json"], [])
            if existing["title"] != item["title"] and "normalized_name_collision" not in risks:
                risks.append("normalized_name_collision")
            status = "draft" if risks else "approved"
            self.db.execute(
                """UPDATE knowledge_graph_nodes SET title=?,is_key=?,is_difficult=?,is_exam=?,notes=?,
                   origin=?,review_status=?,risk_codes_json=?,source_json=?,updated_at=CURRENT_TIMESTAMP
                   WHERE graph_node_id=?""",
                (item["title"], int(item.get("is_key", False)), int(item.get("is_difficult", False)),
                 int(item.get("is_exam", False)), item.get("notes", ""), origin, status,
                 _json(risks), _json(item.get("source") or {}), existing["graph_node_id"]),
            )
            return self.db.fetch_one("SELECT * FROM knowledge_graph_nodes WHERE graph_node_id=?",
                                     (existing["graph_node_id"],)) or {}
        node_id = f"kgn_{uuid.uuid4().hex}"
        self.db.execute(
            """INSERT INTO knowledge_graph_nodes(
                   graph_node_id,graph_id,title,normalized_title,summary,markdown,is_key,is_difficult,
                   is_exam,notes,origin,review_status,risk_codes_json,source_json
               ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (node_id, graph_id, item["title"], normalized, item.get("summary", ""), item.get("markdown", ""),
             int(item.get("is_key", False)), int(item.get("is_difficult", False)),
             int(item.get("is_exam", False)), item.get("notes", ""), origin,
             "approved" if origin == "file" else "draft", "[]", _json(item.get("source") or {})),
        )
        return self.db.fetch_one("SELECT * FROM knowledge_graph_nodes WHERE graph_node_id=?", (node_id,)) or {}

    def _relation_node(self, graph_id: str, title: str, source: dict[str, Any]) -> dict[str, Any]:
        normalized = _normalized_title(title)
        row = self.db.fetch_one(
            "SELECT * FROM knowledge_graph_nodes WHERE graph_id=? AND normalized_title=?", (graph_id, normalized),
        )
        if row:
            return row
        node_id = f"kgn_{uuid.uuid4().hex}"
        self.db.execute(
            """INSERT INTO knowledge_graph_nodes(
                   graph_node_id,graph_id,title,normalized_title,origin,review_status,risk_codes_json,source_json
               ) VALUES(?,?,?,?, 'relation_only','draft',?,?)""",
            (node_id, graph_id, title, normalized, _json(["relation_only_node"]), _json(source)),
        )
        return self.db.fetch_one("SELECT * FROM knowledge_graph_nodes WHERE graph_node_id=?", (node_id,)) or {}

    def commit_import_batch(self, actor: dict[str, Any], batch_id: str) -> dict[str, Any]:
        batch = self._require_batch(actor, batch_id)
        if batch["status"] != "staging":
            return self.get_import_batch(actor, batch_id)
        self.db.execute(
            "UPDATE knowledge_graph_import_batches SET status='committing',updated_at=CURRENT_TIMESTAMP WHERE batch_id=?",
            (batch_id,),
        )
        files = self.db.fetch_all(
            "SELECT * FROM knowledge_graph_import_files WHERE batch_id=? ORDER BY relative_path", (batch_id,),
        )
        parsed_nodes: list[dict[str, Any]] = []
        parsed_relations: list[dict[str, Any]] = []
        definitions: dict[str, str] = {}
        errors = 0
        for row in files:
            if row["confirmed_kind"] == "ignore":
                self.db.execute(
                    "UPDATE knowledge_graph_import_files SET status='ignored',progress=100,stage='已忽略' WHERE file_id=?",
                    (row["file_id"],),
                )
                continue
            try:
                self.db.execute(
                    "UPDATE knowledge_graph_import_files SET status='processing',progress=35,stage='读取工作表' WHERE file_id=?",
                    (row["file_id"],),
                )
                nodes, relations, found_definitions = self._parse_file(row)
                parsed_nodes.extend(nodes)
                parsed_relations.extend(relations)
                definitions.update(found_definitions)
                if not nodes and not relations and not found_definitions:
                    raise ValidationError("未识别到知识点、关系或关系定义表")
                self.db.execute(
                    "UPDATE knowledge_graph_import_files SET status='completed',progress=100,stage='解析完成' WHERE file_id=?",
                    (row["file_id"],),
                )
            except Exception as exc:
                errors += 1
                self.db.execute(
                    """UPDATE knowledge_graph_import_files SET status='failed',progress=100,stage='解析失败',
                       error_message=?,risk_codes_json=? WHERE file_id=?""",
                    (_clean(exc, 500), _json(["parse_failed"]), row["file_id"]),
                )
        graph_id = batch["graph_id"]
        for item in parsed_nodes:
            self._upsert_node(graph_id, item)
        review_count = 0
        inserted_relations = 0
        for item in parsed_relations:
            source = self._relation_node(graph_id, item["source"], item["source_ref"])
            target = self._relation_node(graph_id, item["target"], item["source_ref"])
            mapped = RELATION_MAP.get(item["raw_kind"])
            risks: list[str] = []
            if not mapped:
                mapped = ("related", item["raw_kind"])
                risks.append("unknown_relation_kind")
            kind, label = mapped
            raw_source, raw_target = source, target
            if kind == "prerequisite":
                source, target = target, source
            if kind == "parallel" and source["graph_node_id"] > target["graph_node_id"]:
                source, target = target, source
            if raw_source["origin"] == "relation_only" or raw_target["origin"] == "relation_only":
                risks.append("relation_endpoint_requires_review")
            status = "draft" if risks else "approved"
            relation_id = f"kgr_{uuid.uuid4().hex}"
            before = self.db.fetch_one(
                """SELECT graph_relation_id FROM knowledge_graph_relations
                   WHERE graph_id=? AND source_node_id=? AND target_node_id=? AND relation_kind=?""",
                (graph_id, source["graph_node_id"], target["graph_node_id"], kind),
            )
            self.db.execute(
                """INSERT OR IGNORE INTO knowledge_graph_relations(
                       graph_relation_id,graph_id,source_node_id,target_node_id,relation_kind,relation_label,
                       origin,confidence,reason,review_status,risk_codes_json,source_json
                   ) VALUES(?,?,?,?,?,?, 'explicit',1.0,?,?,?,?)""",
                (relation_id, graph_id, source["graph_node_id"], target["graph_node_id"], kind, label,
                 definitions.get(item["raw_kind"], ""), status, _json(sorted(set(risks))),
                 _json({**item["source_ref"], "raw_source": item["source"],
                        "raw_target": item["target"], "raw_kind": item["raw_kind"]})),
            )
            if not before:
                inserted_relations += 1
                review_count += status == "draft"
        if definitions:
            self.db.execute(
                "UPDATE knowledge_graphs SET relation_definitions_json=?,updated_at=CURRENT_TIMESTAMP WHERE graph_id=?",
                (_json(definitions), graph_id),
            )
        node_count = len({(_normalized_title(item["title"])) for item in parsed_nodes})
        status = "completed_with_warnings" if review_count or errors else "completed"
        self.db.execute(
            """UPDATE knowledge_graph_import_batches SET status=?,node_count=?,relation_count=?,
               review_count=?,error_count=?,updated_at=CURRENT_TIMESTAMP WHERE batch_id=?""",
            (status, node_count, inserted_relations, review_count, errors, batch_id),
        )
        return self.get_import_batch(actor, batch_id)

    @staticmethod
    def _decode_node(row: dict[str, Any]) -> dict[str, Any]:
        value = dict(row)
        value["is_key"] = bool(value["is_key"])
        value["is_difficult"] = bool(value["is_difficult"])
        value["is_exam"] = bool(value["is_exam"])
        value["risk_codes"] = _loads(value.pop("risk_codes_json", "[]"), [])
        value["source"] = _loads(value.pop("source_json", "{}"), {})
        return value

    @staticmethod
    def _decode_relation(row: dict[str, Any]) -> dict[str, Any]:
        value = dict(row)
        value["risk_codes"] = _loads(value.pop("risk_codes_json", "[]"), [])
        value["source"] = _loads(value.pop("source_json", "{}"), {})
        return value

    def workbench(self, actor: dict[str, Any], course_id: str) -> dict[str, Any]:
        graph = self._graph(actor, course_id)
        nodes = [self._decode_node(row) for row in self.db.fetch_all(
            "SELECT * FROM knowledge_graph_nodes WHERE graph_id=? ORDER BY title", (graph["graph_id"],),
        )]
        relations = [self._decode_relation(row) for row in self.db.fetch_all(
            """SELECT r.*,s.title source_title,t.title target_title
               FROM knowledge_graph_relations r JOIN knowledge_graph_nodes s ON s.graph_node_id=r.source_node_id
               JOIN knowledge_graph_nodes t ON t.graph_node_id=r.target_node_id
               WHERE r.graph_id=? ORDER BY r.created_at""", (graph["graph_id"],),
        )]
        versions = self.db.fetch_all(
            "SELECT * FROM knowledge_graph_versions WHERE graph_id=? ORDER BY version_number DESC",
            (graph["graph_id"],),
        )
        batches = self.db.fetch_all(
            """SELECT * FROM knowledge_graph_import_batches WHERE graph_id=?
               ORDER BY created_at DESC LIMIT 20""", (graph["graph_id"],),
        )
        counts = {status: sum(node["review_status"] == status for node in nodes)
                  for status in NODE_STATUSES}
        return {
            "graph": {**graph, "relation_definitions": _loads(graph["relation_definitions_json"], {})},
            "nodes": nodes, "relations": relations, "versions": versions, "batches": batches,
            "metrics": {"nodes": len(nodes), "relations": len(relations),
                        "approved": counts["approved"], "review_required": counts["draft"] + sum(
                            relation["review_status"] == "draft" for relation in relations),
                        "published_version": versions[0]["version_number"] if versions else 0},
        }

    @staticmethod
    def _knowledge_revision(row: dict[str, Any]) -> str:
        payload = _json([row.get("title"), row.get("summary"), row.get("markdown"), row.get("keywords_json")])
        return hashlib.sha256(payload.encode("utf-8")).hexdigest()

    def import_approved_nodes(self, actor: dict[str, Any], course_id: str,
                              node_ids: list[str] | None = None) -> dict[str, Any]:
        graph = self._graph(actor, course_id)
        conditions = ["course_id=?", "status='approved'", "node_type='knowledge_point'",
                      "content_domain='knowledge'"]
        params: list[Any] = [course_id]
        if node_ids:
            unique = list(dict.fromkeys(node_ids))
            conditions.append(f"node_id IN ({','.join('?' for _ in unique)})")
            params.extend(unique)
        rows = self.db.fetch_all(
            f"SELECT * FROM knowledge_nodes WHERE {' AND '.join(conditions)} ORDER BY sort_order", tuple(params),
        )
        if node_ids and len(rows) != len(set(node_ids)):
            raise ValidationError("只能导入当前课程中已审核的知识点")
        imported = 0
        for row in rows:
            revision = self._knowledge_revision(row)
            normalized = _normalized_title(row["title"])
            existing = self.db.fetch_one(
                "SELECT * FROM knowledge_graph_nodes WHERE graph_id=? AND normalized_title=?",
                (graph["graph_id"], normalized),
            )
            source = {"knowledge_node_id": row["node_id"],
                      "source_pages": _loads(row.get("source_pages_json"), [])}
            if existing:
                self.db.execute(
                    """UPDATE knowledge_graph_nodes SET title=?,summary=?,markdown=?,origin='knowledge_center',
                       source_knowledge_node_id=?,source_revision=?,source_json=?,review_status='approved',
                       updated_at=CURRENT_TIMESTAMP WHERE graph_node_id=?""",
                    (row["title"], row["summary"], row["markdown"], row["node_id"], revision,
                     _json(source), existing["graph_node_id"]),
                )
            else:
                self.db.execute(
                    """INSERT INTO knowledge_graph_nodes(
                           graph_node_id,graph_id,title,normalized_title,summary,markdown,origin,
                           source_knowledge_node_id,source_revision,review_status,source_json
                       ) VALUES(?,?,?,?,?,?,'knowledge_center',?,?,'approved',?)""",
                    (f"kgn_{uuid.uuid4().hex}", graph["graph_id"], row["title"], normalized,
                     row["summary"], row["markdown"], row["node_id"], revision, _json(source)),
                )
            imported += 1
        return {"imported": imported, "graph_id": graph["graph_id"]}

    def source_diff(self, actor: dict[str, Any], course_id: str) -> list[dict[str, Any]]:
        graph = self._graph(actor, course_id)
        rows = self.db.fetch_all(
            """SELECT g.graph_node_id,g.title graph_title,g.source_revision,n.*
               FROM knowledge_graph_nodes g LEFT JOIN knowledge_nodes n ON n.node_id=g.source_knowledge_node_id
               WHERE g.graph_id=? AND g.origin='knowledge_center'""", (graph["graph_id"],),
        )
        result = []
        for row in rows:
            state = "source_missing" if not row.get("node_id") else (
                "source_unapproved" if row["status"] != "approved" else (
                    "changed" if self._knowledge_revision(row) != row["source_revision"] else "current"
                )
            )
            result.append({"graph_node_id": row["graph_node_id"], "title": row["graph_title"],
                           "source_knowledge_node_id": row.get("node_id"), "state": state})
        return result

    def sync_sources(self, actor: dict[str, Any], course_id: str,
                     graph_node_ids: list[str] | None = None) -> dict[str, Any]:
        graph = self._graph(actor, course_id)
        conditions = ["g.graph_id=?", "g.origin='knowledge_center'", "n.status='approved'"]
        params: list[Any] = [graph["graph_id"]]
        if graph_node_ids:
            unique = list(dict.fromkeys(graph_node_ids))
            conditions.append(f"g.graph_node_id IN ({','.join('?' for _ in unique)})")
            params.extend(unique)
        rows = self.db.fetch_all(
            f"""SELECT g.graph_node_id,n.* FROM knowledge_graph_nodes g
                 JOIN knowledge_nodes n ON n.node_id=g.source_knowledge_node_id
                 WHERE {' AND '.join(conditions)}""", tuple(params),
        )
        for row in rows:
            self.db.execute(
                """UPDATE knowledge_graph_nodes SET title=?,normalized_title=?,summary=?,markdown=?,
                   source_revision=?,updated_at=CURRENT_TIMESTAMP WHERE graph_node_id=?""",
                (row["title"], _normalized_title(row["title"]), row["summary"], row["markdown"],
                 self._knowledge_revision(row), row["graph_node_id"]),
            )
        return {"synced": len(rows)}

    def update_node(self, actor: dict[str, Any], node_id: str, payload: dict[str, Any]) -> dict[str, Any]:
        row = self.db.fetch_one(
            """SELECT n.*,g.course_id FROM knowledge_graph_nodes n JOIN knowledge_graphs g USING(graph_id)
               JOIN courses c USING(course_id) WHERE n.graph_node_id=? AND c.owner_id=?""",
            (node_id, actor["user_id"]),
        )
        if not row or actor.get("role") != "teacher":
            raise PermissionDenied("无权修改该图谱节点")
        status = str(payload.get("review_status", row["review_status"]))
        if status not in NODE_STATUSES:
            raise ValidationError("节点审核状态无效")
        title = _clean(payload.get("title", row["title"]), 300)
        self.db.execute(
            """UPDATE knowledge_graph_nodes SET title=?,normalized_title=?,summary=?,markdown=?,notes=?,
               is_key=?,is_difficult=?,is_exam=?,review_status=?,updated_at=CURRENT_TIMESTAMP
               WHERE graph_node_id=?""",
            (title, _normalized_title(title), str(payload.get("summary", row["summary"])),
             str(payload.get("markdown", row["markdown"])), _clean(payload.get("notes", row["notes"]), 2000),
             int(payload.get("is_key", row["is_key"])), int(payload.get("is_difficult", row["is_difficult"])),
             int(payload.get("is_exam", row["is_exam"])), status, node_id),
        )
        return self._decode_node(self.db.fetch_one(
            "SELECT * FROM knowledge_graph_nodes WHERE graph_node_id=?", (node_id,)) or {})

    def update_relation(self, actor: dict[str, Any], relation_id: str,
                        payload: dict[str, Any]) -> dict[str, Any]:
        row = self.db.fetch_one(
            """SELECT r.* FROM knowledge_graph_relations r JOIN knowledge_graphs g USING(graph_id)
               JOIN courses c USING(course_id) WHERE r.graph_relation_id=? AND c.owner_id=?""",
            (relation_id, actor["user_id"]),
        )
        if not row or actor.get("role") != "teacher":
            raise PermissionDenied("无权修改该图谱关系")
        status = str(payload.get("review_status", row["review_status"]))
        kind = str(payload.get("relation_kind", row["relation_kind"]))
        if status not in NODE_STATUSES or kind not in RELATION_KINDS:
            raise ValidationError("关系类型或审核状态无效")
        self.db.execute(
            """UPDATE knowledge_graph_relations SET relation_kind=?,relation_label=?,reason=?,
               review_status=?,updated_at=CURRENT_TIMESTAMP WHERE graph_relation_id=?""",
            (kind, _clean(payload.get("relation_label", row["relation_label"]), 120),
             _clean(payload.get("reason", row["reason"]), 1000), status, relation_id),
        )
        return self._decode_relation(self.db.fetch_one(
            "SELECT * FROM knowledge_graph_relations WHERE graph_relation_id=?", (relation_id,)) or {})

    def publish(self, actor: dict[str, Any], course_id: str) -> dict[str, Any]:
        graph = self._graph(actor, course_id)
        nodes = self.db.fetch_all(
            "SELECT * FROM knowledge_graph_nodes WHERE graph_id=? AND review_status='approved' ORDER BY title",
            (graph["graph_id"],),
        )
        if not nodes:
            raise ValidationError("没有已批准的图谱节点")
        node_ids = {row["graph_node_id"] for row in nodes}
        relations = [row for row in self.db.fetch_all(
            "SELECT * FROM knowledge_graph_relations WHERE graph_id=? AND review_status='approved'",
            (graph["graph_id"],),
        ) if row["source_node_id"] in node_ids and row["target_node_id"] in node_ids]
        current = self.db.fetch_one(
            "SELECT COALESCE(MAX(version_number),0) n FROM knowledge_graph_versions WHERE graph_id=?",
            (graph["graph_id"],),
        ) or {"n": 0}
        version_id = f"kgv_{uuid.uuid4().hex}"
        with self.db.connect() as conn:
            conn.execute(
                "UPDATE knowledge_graph_versions SET status='superseded' WHERE graph_id=? AND status='published'",
                (graph["graph_id"],),
            )
            conn.execute(
                """INSERT INTO knowledge_graph_versions(
                       graph_version_id,graph_id,course_id,version_number,node_count,relation_count,created_by
                   ) VALUES(?,?,?,?,?,?,?)""",
                (version_id, graph["graph_id"], course_id, int(current["n"]) + 1,
                 len(nodes), len(relations), actor["user_id"]),
            )
            conn.executemany(
                "INSERT INTO knowledge_graph_version_nodes(graph_version_id,graph_node_id,snapshot_json) VALUES(?,?,?)",
                [(version_id, row["graph_node_id"], _json(self._decode_node(row))) for row in nodes],
            )
            conn.executemany(
                "INSERT INTO knowledge_graph_version_relations(graph_version_id,graph_relation_id,snapshot_json) VALUES(?,?,?)",
                [(version_id, row["graph_relation_id"], _json(self._decode_relation(row))) for row in relations],
            )
        return self.db.fetch_one(
            "SELECT * FROM knowledge_graph_versions WHERE graph_version_id=?", (version_id,)) or {}

    def student_graph(self, actor: dict[str, Any], course_id: str) -> dict[str, Any]:
        if actor.get("role") != "student":
            raise PermissionDenied("仅学生可以查看已发布课程知识图谱")
        self.campus.require_access(course_id, str(actor["user_id"]), "student")
        version = self.db.fetch_one(
            """SELECT v.* FROM knowledge_graph_versions v JOIN knowledge_graphs g USING(graph_id)
               WHERE g.course_id=? AND v.status='published' ORDER BY v.version_number DESC LIMIT 1""",
            (course_id,),
        )
        if not version:
            raise NotFound("该课程尚未发布知识图谱")
        nodes = [_loads(row["snapshot_json"], {}) for row in self.db.fetch_all(
            "SELECT snapshot_json FROM knowledge_graph_version_nodes WHERE graph_version_id=?",
            (version["graph_version_id"],),
        )]
        relations = [_loads(row["snapshot_json"], {}) for row in self.db.fetch_all(
            "SELECT snapshot_json FROM knowledge_graph_version_relations WHERE graph_version_id=?",
            (version["graph_version_id"],),
        )]
        for node in nodes:
            node.pop("source_revision", None)
            node.pop("risk_codes", None)
            source = node.get("source") or {}
            node["source"] = {key: source[key] for key in ("file", "sheet", "row", "source_pages") if key in source}
            if node.get("origin") == "knowledge_center":
                node["markdown"] = ""
        return {"version": version, "nodes": nodes, "relations": relations}
