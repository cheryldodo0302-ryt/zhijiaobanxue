import io

import pytest
from openpyxl import Workbook

from auth_service import AuthService
from campus_service import CampusService, PermissionDenied
from database import LearningDatabase
from ingestion_service import IngestionService
from question_bank_service import QuestionBankService
from teacher_service import TeacherService


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def question_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append([
        "题目类型", "答案类型", "题干内容", "答案", "答案解析",
        "主题词一", "知识点", "难度", "答题时间（秒）",
        "选项A", "选项B", "选项C", "选项D",
    ])
    sheet.append([
        "判断题", "判断题", "关系模型中的元组可以重复。", "错", "关系要求元组唯一。",
        "关系数据库", "关系模型", "中等", 15, None, None, None, None,
    ])
    sheet.append([
        "选择题", "单选题", "SQL 查询的核心关键字是？", "A", "SELECT 用于查询。",
        "数据库语言", "SQL", "容易", 20, "SELECT", "UPDATE", "DELETE", "INSERT",
    ])
    sheet.append([
        "选择题", "单选题", "缺少答案的无效题", None, None,
        "数据库语言", "SQL", "容易", 20, "A", "B", None, None,
    ])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


@pytest.fixture()
def question_bank(tmp_path):
    db = LearningDatabase(tmp_path / "question-bank.db")
    campus = CampusService(db, tmp_path / "uploads", provider_factory=lambda: None)
    auth = AuthService(db, tmp_path / "secret")
    teacher = auth.create_user("bank_teacher", "safe-password-123", "teacher")
    other_teacher = auth.create_user("other_teacher", "safe-password-456", "teacher")
    student_one = auth.create_user("student_one", "safe-password-789", "student")
    student_two = auth.create_user("student_two", "safe-password-012", "student")
    course = TeacherService(db, campus).create_course(teacher, "数据库原理")
    campus.enroll_student(course["course_id"], teacher["user_id"], student_one["user_id"])
    campus.enroll_student(course["course_id"], teacher["user_id"], student_two["user_id"])
    return (
        db, campus, IngestionService(db, campus), QuestionBankService(db, campus),
        teacher, other_teacher, student_one, student_two, course,
    )


def test_template_import_review_publish_answer_and_statistics(question_bank):
    db, _, ingestion, service, teacher, _, student_one, student_two, course = question_bank
    content = question_workbook()
    imported = service.import_template(
        teacher, course["course_id"], "题库模板.xlsx", XLSX_MIME, content
    )
    assert imported["valid_rows"] == 3
    assert imported["invalid_rows"] == 0
    assert any(warning["row"] == 4 and "答案" in warning["message"]
               for warning in imported["warnings"])
    assert service.import_template(
        teacher, course["course_id"], "题库模板.xlsx", XLSX_MIME, content
    )["duplicate"] is True

    items = ingestion.list_question_bank(teacher, course["course_id"])
    assert {item["source_kind"] for item in items} == {"teacher_template"}
    judgment = next(item for item in items if item["question_type"] == "true_false")
    choice = next(item for item in items if "SQL 查询" in item["stem_markdown"])
    assert judgment["answer_markdown"] == "F"
    assert choice["options"][0] == {"key": "A", "text": "SELECT"}
    for item in items:
        if not item["answer_markdown"]:
            continue
        ingestion.review_question(teacher, item["item_id"], {"status": "approved"})
    version = ingestion.publish_question_bank(teacher, course["course_id"])

    student_actor = {"user_id": student_one["user_id"], "role": "student"}
    published = service.student_questions(student_actor, course["course_id"])
    assert published["version_id"] == version["version_id"]
    assert len(published["items"]) == 2
    assert "correct_answer" not in published["items"][0]

    first_result = service.submit(student_actor, course["course_id"], version["version_id"], [
        {"item_id": published["items"][0]["item_id"], "response": "对"},
        {"item_id": published["items"][1]["item_id"], "response": "A"},
    ])
    assert first_result["accuracy"] == 50.0
    service.submit(
        {"user_id": student_two["user_id"], "role": "student"},
        course["course_id"], version["version_id"], [
            {"item_id": published["items"][0]["item_id"], "response": "对"},
            {"item_id": published["items"][1]["item_id"], "response": "B"},
        ],
    )
    statistics = service.statistics(teacher, course["course_id"])
    assert statistics["summary"]["students"] == 2
    assert statistics["summary"]["attempts"] == 4
    assert statistics["ranking"][0]["error_rate"] == 100.0
    assert all(student["wrong_questions"] for student in statistics["students"])
    assert db.fetch_one("SELECT COUNT(*) count FROM question_bank_attempts")["count"] == 4


def test_true_false_answers_use_tf_but_accept_legacy_inputs(question_bank):
    _, _, _, service, _, _, _, _, _ = question_bank
    assert service._normalize_true_false("Y") == "T"
    assert service._normalize_true_false("N") == "F"
    assert service._normalize_true_false("对") == "T"
    assert service._normalize_true_false("错误") == "F"
    assert service._is_correct("true_false", "Y", "T") is True
    assert service._is_correct("true_false", "错", "F") is True


def test_folder_import_move_publish_and_student_listing(question_bank):
    _, _, ingestion, service, teacher, _, student_one, _, course = question_bank
    folder = service.create_folder(teacher, course["course_id"], "第三章作业", "homework")
    imported = service.import_template(
        teacher, course["course_id"], "第三章.xlsx", XLSX_MIME, question_workbook(),
        ai_mode="local", folder_id=folder["folder_id"],
    )
    assert imported["valid_rows"] == 3
    items = ingestion.list_question_bank(teacher, course["course_id"])
    assert {item["folder_id"] for item in items} == {folder["folder_id"]}
    assert not items[0]["answer_markdown"]  # 异常题置顶
    for item in items:
        if item["answer_markdown"]:
            ingestion.review_question(teacher, item["item_id"], {"status": "approved"})
    version = ingestion.publish_question_bank(
        teacher, course["course_id"], folder_id=folder["folder_id"]
    )
    student = {"user_id": student_one["user_id"], "role": "student"}
    publications = service.student_publications(student, course["course_id"])
    assert publications[0]["folder_name"] == "第三章作业"
    published = service.student_questions(
        student, course["course_id"], folder_id=folder["folder_id"]
    )
    assert published["version_id"] == version["version_id"]
    assert published["folder_id"] == folder["folder_id"]


def test_judgment_keeps_excel_yn_options(question_bank):
    _, _, ingestion, service, teacher, _, _, _, course = question_bank
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["题目类型", "题干", "正确答案", "选项A", "选项B"])
    sheet.append(["判断题", "事务具有原子性。", "Y", "Y", "N"])
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    service.import_template(
        teacher, course["course_id"], "YN判断题.xlsx", XLSX_MIME,
        output.getvalue(), ai_mode="local",
    )
    item = ingestion.list_question_bank(teacher, course["course_id"])[0]
    assert item["answer_markdown"] == "Y"
    assert [option["text"] for option in item["options"]] == ["Y", "N"]
    assert service._is_correct("true_false", "Y", item["answer_markdown"]) is True


def test_question_bank_permissions_and_textbook_examples_do_not_create_formal_items(question_bank):
    _, _, ingestion, service, teacher, other_teacher, _, _, course = question_bank
    with pytest.raises(PermissionDenied):
        service.import_template(
            other_teacher, course["course_id"], "题库模板.xlsx", XLSX_MIME, question_workbook()
        )
    document = ingestion.queue_document(
        teacher, course["course_id"], "教材例题.txt", "text/plain", "例题：1+1 等于几？".encode()
    )
    ingestion.process_job(document["job_id"])
    block = ingestion.list_blocks(teacher, document["document_id"])[0]
    ingestion.update_classification(
        teacher, block["block_id"], destination="question_bank",
        semantic_role="example", question_group_key="example-1", reason="教材例题",
    )
    assert ingestion.list_question_bank(teacher, course["course_id"]) == []


def test_shifted_alias_headers_are_detected_without_fixed_columns(question_bank):
    _, _, ingestion, service, teacher, _, _, _, course = question_bank
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "自由格式"
    sheet.append(["数据库原理自建题库"])
    sheet.append([])
    sheet.append(["编号", "章节", "试题", "B", "正确答案", "A", "说明", "难易度"])
    sheet.append([1, "SQL", "用于查询数据的关键字是？", "UPDATE", "A", "SELECT", "查询使用 SELECT", "容易"])
    output = io.BytesIO(); workbook.save(output); workbook.close()

    imported = service.import_template(
        teacher, course["course_id"], "自由列题库.xlsx", XLSX_MIME, output.getvalue(), ai_mode="local"
    )
    assert imported["valid_rows"] == 1
    assert imported["detected_schema"][0]["header_row"] == 3
    item = ingestion.list_question_bank(teacher, course["course_id"])[0]
    assert item["stem_markdown"] == "用于查询数据的关键字是？"
    assert item["options"] == [{"key": "A", "text": "SELECT"}, {"key": "B", "text": "UPDATE"}]


class FlexibleBankProvider:
    def generate_json(self, _system, _prompt):
        return {"questions": [{
            "source_sheet": "非标准", "source_row": 1, "type": "single_choice",
            "stem": "事务的四个特性缩写是？",
            "options": {"A": "ACID", "B": "BASE"}, "answer": "A",
            "explanation": "原单元格注明答案 A。", "knowledge_points": ["事务"],
            "difficulty": "中等", "confidence": 0.91, "notes": [],
        }]}


def test_ai_can_structure_unrecognized_content_without_becoming_mandatory(question_bank):
    _, campus, ingestion, service, teacher, _, _, _, course = question_bank
    campus.provider_factory = FlexibleBankProvider
    workbook = Workbook(); sheet = workbook.active; sheet.title = "非标准"
    sheet.append(["第1题 事务的四个特性缩写是？ A.ACID B.BASE 答案A"])
    output = io.BytesIO(); workbook.save(output); workbook.close()

    imported = service.import_template(
        teacher, course["course_id"], "非标准题库.xlsx", XLSX_MIME, output.getvalue(), ai_mode="auto"
    )
    assert imported["ai_used"] is True
    item = ingestion.list_question_bank(teacher, course["course_id"])[0]
    assert item["recognition_method"] == "ai"
    assert item["answer_markdown"] == "A"
