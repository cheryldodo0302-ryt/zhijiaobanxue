from __future__ import annotations

import hashlib
import io
import json
import re
import uuid
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from campus_service import CampusService, NotFound, PermissionDenied, ValidationError, _safe_name
from database import LearningDatabase
from llm_provider import GeminiProvider, QwenProvider


QUESTION_BANK_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
    "application/xls",
    "application/x-xls",
    "application/octet-stream",
}
OPTION_LETTERS = tuple("ABCDEFGHIJKLMNO")
FIELD_ALIASES = {
    "stem": ("题干内容", "题干", "题目内容", "试题内容", "题目", "试题", "问题", "question", "stem"),
    "answer": ("标准答案", "正确答案", "参考答案", "答案", "answer", "key"),
    "explanation": ("答案解析", "试题解析", "解析", "说明", "解题说明", "analysis", "explanation"),
    "question_type": ("题目类型", "试题类型", "题型", "类型", "questiontype", "type"),
    "answer_type": ("答案类型", "作答类型", "answertype"),
    "difficulty": ("难度", "难易度", "difficulty"),
    "duration": ("答题时间秒", "建议用时秒", "答题时间", "用时", "duration", "time"),
    "knowledge": ("知识点", "考点", "章节", "knowledgepoint", "knowledge"),
    "topic1": ("主题词一", "主题词1", "关键词一", "关键词1", "标签一"),
    "topic2": ("主题词二", "主题词2", "关键词二", "关键词2", "标签二"),
    "topic3": ("主题词三", "主题词3", "关键词三", "关键词3", "标签三"),
}


def _json(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False)


def _loads(value: str | None, default: Any) -> Any:
    try:
        return json.loads(value or "")
    except (TypeError, ValueError, json.JSONDecodeError):
        return default


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


class QuestionBankService:
    """Reviewed, versioned question bank imported by a course teacher."""

    def __init__(self, db: LearningDatabase, campus: CampusService):
        self.db = db
        self.campus = campus

    def _teacher_course(self, actor: dict[str, Any], course_id: str) -> dict[str, Any]:
        if actor.get("role") != "teacher":
            raise PermissionDenied("仅教师可以管理共享题库")
        course = self.campus.require_access(course_id, str(actor["user_id"]), "teacher")
        if course["course_type"] != "shared_course" or course["owner_id"] != actor["user_id"]:
            raise PermissionDenied("只能管理自己创建的共享课程题库")
        return course

    def _folder(self, actor: dict[str, Any], course_id: str,
                folder_id: str | None) -> dict[str, Any] | None:
        self._teacher_course(actor, course_id)
        if not folder_id:
            return None
        folder = self.db.fetch_one(
            "SELECT * FROM question_bank_folders WHERE folder_id=? AND course_id=?",
            (folder_id, course_id),
        )
        if not folder:
            raise ValidationError("题库文件夹不存在或不属于当前课程")
        return folder

    def create_folder(self, actor: dict[str, Any], course_id: str, name: str,
                      folder_type: str, *, parent_folder_id: str | None = None,
                      relative_path: str = "") -> dict[str, Any]:
        self._teacher_course(actor, course_id)
        name = _text(name)
        if not name:
            raise ValidationError("文件夹名称不能为空")
        if folder_type not in {"exam", "homework", "chapter"}:
            raise ValidationError("文件夹类型必须是试卷、作业或章节练习")
        if parent_folder_id:
            self._folder(actor, course_id, parent_folder_id)
        clean_path = "/".join(part for part in str(relative_path or "").replace("\\", "/").split("/") if part)
        if any(part in {".", ".."} for part in clean_path.split("/") if part) or len(clean_path) > 500:
            raise ValidationError("题库目录路径不安全")
        folder_id = f"qbf_{uuid.uuid4().hex}"
        try:
            self.db.execute(
                """INSERT INTO question_bank_folders(
                       folder_id,course_id,folder_name,folder_type,created_by,parent_folder_id,relative_path
                   ) VALUES(?,?,?,?,?,?,?)""",
                (folder_id, course_id, name[:120], folder_type, actor["user_id"], parent_folder_id, clean_path),
            )
        except Exception as exc:
            if "UNIQUE" in str(exc).upper():
                raise ValidationError("当前课程已有同名题库文件夹") from exc
            raise
        return self.db.fetch_one(
            "SELECT * FROM question_bank_folders WHERE folder_id=?", (folder_id,)
        ) or {}

    def list_folders(self, actor: dict[str, Any], course_id: str) -> list[dict[str, Any]]:
        self._teacher_course(actor, course_id)
        return self.db.fetch_all(
            """SELECT f.*,COUNT(q.item_id) item_count,
                      SUM(CASE WHEN q.status='draft' THEN 1 ELSE 0 END) draft_count,
                      SUM(CASE WHEN q.status='approved' THEN 1 ELSE 0 END) approved_count,
                      SUM(CASE WHEN q.status='draft' AND (
                          q.answer_markdown='' OR q.recognition_confidence<0.7
                          OR q.recognition_notes_json NOT IN ('','[]')
                      ) THEN 1 ELSE 0 END) issue_count
               FROM question_bank_folders f
               LEFT JOIN question_bank_items q ON q.folder_id=f.folder_id
               WHERE f.course_id=? GROUP BY f.folder_id ORDER BY f.created_at""",
            (course_id,),
        )

    def move_items(self, actor: dict[str, Any], course_id: str,
                   item_ids: list[str], folder_id: str | None) -> dict[str, Any]:
        self._folder(actor, course_id, folder_id)
        unique_ids = list(dict.fromkeys(_text(value) for value in item_ids if _text(value)))
        if not unique_ids or len(unique_ids) > 500:
            raise ValidationError("请选择 1 至 500 道题")
        placeholders = ",".join("?" for _ in unique_ids)
        found = self.db.fetch_all(
            f"SELECT item_id FROM question_bank_items WHERE course_id=? AND item_id IN ({placeholders})",
            (course_id, *unique_ids),
        )
        if len(found) != len(unique_ids):
            raise ValidationError("所选题目包含不存在或不属于当前课程的题目")
        self.db.execute(
            f"""UPDATE question_bank_items SET folder_id=?,updated_at=CURRENT_TIMESTAMP
                WHERE course_id=? AND item_id IN ({placeholders})""",
            (folder_id, course_id, *unique_ids),
        )
        return {"moved": len(unique_ids), "folder_id": folder_id}

    @staticmethod
    def _question_type(raw_type: str, raw_answer_type: str, answer: str) -> str:
        value = f"{raw_type} {raw_answer_type}".lower()
        if any(token in value for token in ("判断", "是非", "true", "boolean")):
            return "true_false"
        if "多选" in value or len(re.findall(r"[A-O]", answer.upper())) > 1:
            return "multiple_choice"
        if any(token in value for token in ("选择", "单选", "choice")):
            return "single_choice"
        if any(token in value for token in ("简答", "问答", "short")):
            return "short_answer"
        return "other"

    @staticmethod
    def _normalize_true_false(value: str) -> str:
        compact = re.sub(r"\s+", "", value).lower()
        if compact in {"对", "正确", "是", "√", "true", "t", "y", "yes", "1"}:
            return "T"
        if compact in {"错", "错误", "否", "×", "false", "f", "n", "no", "0"}:
            return "F"
        return ""

    @staticmethod
    def _choice_answers(value: str) -> list[str]:
        letters = re.findall(r"[A-O]", value.upper())
        return list(dict.fromkeys(letters))

    @staticmethod
    def _header_key(value: Any) -> str:
        return re.sub(r"[^a-z0-9\u4e00-\u9fff]", "", _text(value).lower())

    @classmethod
    def _canonical_header(cls, value: Any) -> tuple[str, str] | None:
        key = cls._header_key(value)
        if not key:
            return None
        for field, aliases in FIELD_ALIASES.items():
            if key in {cls._header_key(alias) for alias in aliases}:
                return "field", field
        option = re.fullmatch(r"(?:选项|option)?([a-o])(?:选项)?", key)
        if option:
            return "option", option.group(1).upper()
        numbered = re.fullmatch(r"(?:选项|option)(1[0-5]|[1-9])", key)
        if numbered:
            return "option", OPTION_LETTERS[int(numbered.group(1)) - 1]
        return None

    @staticmethod
    def _read_sheets(file_name: str, data: bytes) -> list[tuple[str, list[list[Any]]]]:
        suffix = Path(file_name).suffix.lower()
        if suffix == ".xlsx":
            try:
                workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
            except Exception as exc:
                raise ValidationError("Excel 文件损坏或不是有效的 XLSX 文件") from exc
            try:
                return [(sheet.title, [list(row) for row in sheet.iter_rows(values_only=True)])
                        for sheet in workbook.worksheets]
            finally:
                workbook.close()
        if suffix == ".xls":
            try:
                import xlrd
                workbook = xlrd.open_workbook(file_contents=data, on_demand=True)
            except ImportError as exc:
                raise ValidationError("服务器缺少 XLS 解析组件 xlrd，请安装项目依赖后重试") from exc
            except Exception as exc:
                raise ValidationError("Excel 文件损坏或不是有效的 XLS 文件") from exc
            try:
                return [(sheet.name, [sheet.row_values(index) for index in range(sheet.nrows)])
                        for sheet in (workbook.sheet_by_index(i) for i in range(workbook.nsheets))]
            finally:
                workbook.release_resources()
        raise ValidationError("题库文件仅支持 .xls 或 .xlsx")

    @classmethod
    def _detect_header(cls, rows: list[list[Any]]) -> tuple[int | None, dict[str, int], dict[str, int], int]:
        best: tuple[int | None, dict[str, int], dict[str, int], int] = (None, {}, {}, 0)
        weights = {"stem": 7, "answer": 5, "question_type": 3, "answer_type": 2,
                   "explanation": 2, "knowledge": 1, "difficulty": 1, "duration": 1}
        for row_index, row in enumerate(rows[:30]):
            fields: dict[str, int] = {}
            options: dict[str, int] = {}
            for column_index, value in enumerate(row):
                matched = cls._canonical_header(value)
                if matched:
                    kind, name = matched
                    (options if kind == "option" else fields).setdefault(name, column_index)
            score = sum(weights.get(name, 1) for name in fields) + min(len(options), 4)
            if score > best[3]:
                best = row_index, fields, options, score
        if best[3] < 5 or ("stem" not in best[1] and "answer" not in best[1]):
            return None, {}, {}, best[3]
        return best

    def _parse_flexible_row(self, values: dict[str, Any], row_number: int, sheet_name: str,
                            option_values: dict[str, Any]) -> tuple[dict[str, Any] | None, list[str]]:
        issues: list[str] = []
        stem, raw_answer = _text(values.get("stem")), _text(values.get("answer"))
        if not stem:
            return None, ["没有识别到题干"]
        options = [{"key": letter, "text": _text(option_values.get(letter))}
                   for letter in OPTION_LETTERS if _text(option_values.get(letter))]
        question_type = self._question_type(
            _text(values.get("question_type")), _text(values.get("answer_type")), raw_answer)
        if question_type == "other" and len(options) >= 2:
            question_type = "single_choice"
        answer: str | list[str] = raw_answer
        if not raw_answer:
            issues.append("未识别到显式标准答案，请教师补充后再批准")
        elif question_type == "true_false":
            compact_answer = re.sub(r"\s+", "", raw_answer).upper()
            answer = compact_answer if compact_answer in {"Y", "N", "T", "F"} else (
                self._normalize_true_false(raw_answer) or raw_answer
            )
            if answer == raw_answer and not self._normalize_true_false(raw_answer):
                issues.append("判断题答案无法自动归一化")
        elif question_type in {"single_choice", "multiple_choice"}:
            keys = self._choice_answers(raw_answer)
            if not keys:
                keys = [option["key"] for option in options if option["text"] == raw_answer]
            if keys:
                answer = keys if question_type == "multiple_choice" else keys[0]
            else:
                issues.append("选择题答案未能映射到选项")
            if len(options) < 2:
                issues.append("选择题未识别到至少两个选项")
        duration = None
        if values.get("duration") not in (None, ""):
            try:
                duration = max(1, min(86400, int(float(values["duration"]))))
            except (TypeError, ValueError):
                issues.append("答题时间不是有效秒数")
        points = list(dict.fromkeys(filter(None, (
            _text(values.get("topic1")), _text(values.get("topic2")),
            _text(values.get("topic3")), _text(values.get("knowledge")),
        ))))
        return {
            "sheet_name": sheet_name, "row_number": row_number, "question_type": question_type,
            "stem": stem, "answer": answer,
            "answer_markdown": ",".join(answer) if isinstance(answer, list) else _text(answer),
            "explanation": _text(values.get("explanation")), "options": options,
            "knowledge_points": points, "difficulty": _text(values.get("difficulty")),
            "duration_seconds": duration, "recognition_method": "local",
            "recognition_confidence": max(0.35, 0.98 - 0.13 * len(issues)),
            "recognition_notes": issues,
        }, issues

    def _local_flexible_parse(self, sheets: list[tuple[str, list[list[Any]]]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], int]:
        parsed: list[dict[str, Any]] = []
        unresolved: list[dict[str, Any]] = []
        schemas: list[dict[str, Any]] = []
        candidate_count = 0
        for sheet_name, rows in sheets:
            header_index, fields, option_columns, score = self._detect_header(rows)
            schemas.append({"sheet": sheet_name,
                            "header_row": None if header_index is None else header_index + 1,
                            "fields": sorted(fields), "options": sorted(option_columns), "score": score})
            start = header_index + 1 if header_index is not None else 0
            for row_index, row in enumerate(rows[start:], start + 1):
                cells = [_text(value) for value in row]
                if not any(cells):
                    continue
                if header_index is None:
                    meaningful = [value for value in cells if value]
                    if len(" ".join(meaningful)) >= 8:
                        unresolved.append({"sheet": sheet_name, "row": row_index, "cells": meaningful[:40]})
                        candidate_count += 1
                    continue
                values = {name: row[column] if column < len(row) else None
                          for name, column in fields.items()}
                option_values = {letter: row[column] if column < len(row) else None
                                 for letter, column in option_columns.items()}
                item, issues = self._parse_flexible_row(values, row_index, sheet_name, option_values)
                candidate_count += 1
                record = {"sheet": sheet_name, "row": row_index, "cells": cells[:40]}
                if item is None:
                    unresolved.append(record)
                else:
                    parsed.append(item)
                    if issues:
                        record["local_item"] = item
                        unresolved.append(record)
        return parsed, unresolved, schemas, candidate_count

    def _provider_for_import(self, ai_mode: str, ai_settings: dict[str, str] | None):
        if ai_mode == "local":
            return None
        settings = ai_settings or {}
        api_key = _text(settings.get("api_key"))
        if api_key:
            base_url, model = _text(settings.get("base_url")), _text(settings.get("model"))
            provider_name = _text(settings.get("provider") or "openai_compatible").lower()
            if not base_url or not model:
                raise ValidationError("使用教师自有 API 时必须填写 Base URL 和模型名称")
            if provider_name in {"gemini", "google", "google_gemini"}:
                return GeminiProvider(api_key, base_url, model, timeout=75)
            return QwenProvider(api_key, base_url, model, timeout=75)
        try:
            return self.campus.provider_factory()
        except Exception:
            return None

    def _normalize_ai_question(self, raw: dict[str, Any], fallback: dict[str, Any]) -> dict[str, Any] | None:
        stem = _text(raw.get("stem") or raw.get("question"))
        if not stem:
            return None
        raw_options = raw.get("options") or []
        options: list[dict[str, str]] = []
        if isinstance(raw_options, dict):
            options = [{"key": _text(key).upper(), "text": _text(value)}
                       for key, value in raw_options.items() if _text(value)]
        elif isinstance(raw_options, list):
            for index, option in enumerate(raw_options[:15]):
                if isinstance(option, dict):
                    key = _text(option.get("key") or option.get("label") or OPTION_LETTERS[index]).upper()
                    text = _text(option.get("text") or option.get("value"))
                else:
                    matched = re.match(r"^\s*([A-O])[.、:：)）\s]+(.*)$", _text(option), re.I)
                    key = matched.group(1).upper() if matched else OPTION_LETTERS[index]
                    text = matched.group(2).strip() if matched else _text(option)
                if text:
                    options.append({"key": key, "text": text})
        raw_answer = raw.get("answer", "")
        answer_text = ",".join(map(str, raw_answer)) if isinstance(raw_answer, list) else _text(raw_answer)
        question_type = self._question_type(_text(raw.get("type")), "", answer_text)
        if question_type == "other" and len(options) >= 2:
            question_type = "single_choice"
        answer: str | list[str] = raw_answer if isinstance(raw_answer, list) else answer_text
        if question_type == "true_false" and answer_text:
            compact_answer = re.sub(r"\s+", "", answer_text).upper()
            answer = compact_answer if compact_answer in {"Y", "N", "T", "F"} else (
                self._normalize_true_false(answer_text) or answer_text
            )
        elif question_type in {"single_choice", "multiple_choice"} and answer_text:
            keys = self._choice_answers(answer_text)
            if keys:
                answer = keys if question_type == "multiple_choice" else keys[0]
        notes = [_text(value) for value in raw.get("notes", []) if _text(value)] \
            if isinstance(raw.get("notes"), list) else []
        if not answer_text:
            notes.append("AI 未在原文件中找到显式答案，请教师补充")
        try:
            confidence = max(0.0, min(1.0, float(raw.get("confidence", 0.65))))
        except (TypeError, ValueError):
            confidence = 0.65
        return {
            "sheet_name": _text(raw.get("source_sheet") or fallback.get("sheet")),
            "row_number": int(raw.get("source_row") or fallback.get("row") or 0),
            "question_type": question_type, "stem": stem, "answer": answer,
            "answer_markdown": ",".join(map(str, answer)) if isinstance(answer, list) else _text(answer),
            "explanation": _text(raw.get("explanation")), "options": options,
            "knowledge_points": [str(value).strip() for value in (raw.get("knowledge_points") or [])
                                 if str(value).strip()],
            "difficulty": _text(raw.get("difficulty")), "duration_seconds": None,
            "recognition_method": "ai", "recognition_confidence": confidence,
            "recognition_notes": list(dict.fromkeys(notes)),
        }

    def _ai_parse_records(self, provider: Any, records: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[str]]:
        if provider is None or not records:
            return [], []
        output: list[dict[str, Any]] = []
        warnings: list[str] = []
        consecutive_connection_failures = 0
        system = (
            "你是高校题库导入整理器。只提取输入单元格中明确存在的信息，不得补写题目、猜测答案或使用外部知识。"
            "识别题干、题型、选项、显式答案、解析、知识点，并保留 source_sheet 与 source_row。"
            "输出 JSON 对象，根字段为 questions。每题字段为 source_sheet、source_row、type、stem、"
            "options（key/text 数组）、answer、explanation、knowledge_points、difficulty、confidence、notes。"
            "type 只能是 single_choice、multiple_choice、true_false、short_answer、other。"
            "没有明确答案时 answer 必须为空。"
        )
        for start in range(0, len(records), 12):
            batch = records[start:start + 12]
            compact = [{"source_sheet": row["sheet"], "source_row": row["row"], "cells": row["cells"]}
                       for row in batch]
            try:
                if hasattr(provider, "generate_json"):
                    result = provider.generate_json(system, _json({"rows": compact}))
                else:
                    raw_text = provider.generate(system, _json({"rows": compact}))
                    cleaned = re.sub(r"^```(?:json)?\s*|\s*```$", "", raw_text.strip(), flags=re.I | re.S)
                    result = json.loads(cleaned)
                questions = result.get("questions", []) if isinstance(result, dict) else result
                for raw_question in questions if isinstance(questions, list) else []:
                    if not isinstance(raw_question, dict):
                        continue
                    source_row = int(raw_question.get("source_row") or 0)
                    source_sheet = _text(raw_question.get("source_sheet"))
                    fallback = next((row for row in batch if row["row"] == source_row and
                                     (not source_sheet or row["sheet"] == source_sheet)), batch[0])
                    normalized = self._normalize_ai_question(raw_question, fallback)
                    if normalized:
                        output.append(normalized)
                consecutive_connection_failures = 0
            except Exception as exc:
                message = str(exc)
                warnings.append(f"AI 整理第 {start // 12 + 1} 批失败，本地结果已保留：{message[:240]}")
                lowered = message.lower()
                if any(marker in lowered for marker in (
                    "timeout", "timed out", "connectionerror", "无法连接智能服务",
                    "max retries exceeded",
                )):
                    consecutive_connection_failures += 1
                    if consecutive_connection_failures >= 2:
                        remaining = max(0, (len(records) - start - len(batch) + 11) // 12)
                        if remaining:
                            warnings.append(
                                f"智能服务连续连接失败，已停止后续 {remaining} 批调用；"
                                "请改用教师自有 API 后重新导入"
                            )
                        break
                else:
                    consecutive_connection_failures = 0
        return output, warnings

    @staticmethod
    def _merge_recognized_items(local_items: list[dict[str, Any]],
                                ai_items: list[dict[str, Any]]) -> list[dict[str, Any]]:
        indexed = {(item["sheet_name"], item["row_number"]): item for item in local_items}
        for ai_item in ai_items:
            key = (ai_item["sheet_name"], ai_item["row_number"])
            local = indexed.get(key)
            if local is None:
                indexed[key] = ai_item
                continue
            for field in ("stem", "answer", "answer_markdown", "explanation", "options",
                          "knowledge_points", "difficulty", "duration_seconds"):
                if not local.get(field) and ai_item.get(field):
                    local[field] = ai_item[field]
            local["recognition_notes"] = list(dict.fromkeys(
                [*local.get("recognition_notes", []), *ai_item.get("recognition_notes", [])]
            ))
            local["recognition_method"] = "local+ai"
            local["recognition_confidence"] = max(
                float(local.get("recognition_confidence", 0)), float(ai_item.get("recognition_confidence", 0)))
        return list(indexed.values())

    def import_flexible(self, actor: dict[str, Any], course_id: str, file_name: str,
                        mime_type: str, data: bytes, *, ai_mode: str = "auto",
                        ai_settings: dict[str, str] | None = None,
                        folder_id: str | None = None) -> dict[str, Any]:
        self._folder(actor, course_id, folder_id)
        suffix = Path(file_name).suffix.lower()
        if suffix not in {".xls", ".xlsx"}:
            raise ValidationError("题库文件支持 Excel .xls 和 .xlsx")
        if not data or len(data) > 20 * 1024 * 1024:
            raise ValidationError("题库文件不能为空且不能超过 20MB")
        if suffix == ".xlsx" and not data.startswith(b"PK"):
            raise ValidationError("文件扩展名为 .xlsx，但内容不是有效的 XLSX 文件")
        if suffix == ".xls" and not data.startswith(bytes.fromhex("D0CF11E0")):
            raise ValidationError("文件扩展名为 .xls，但内容不是有效的 XLS 文件")
        if ai_mode not in {"auto", "local"}:
            raise ValidationError("AI 识别模式无效")
        digest = hashlib.sha256(data).hexdigest()
        existing = self.db.fetch_one(
            "SELECT * FROM question_bank_imports WHERE course_id=? AND sha256=?", (course_id, digest))
        if existing:
            return {**self._format_import(existing), "duplicate": True}

        local_items, unresolved, schemas, total_rows = self._local_flexible_parse(
            self._read_sheets(file_name, data))
        provider = self._provider_for_import(ai_mode, ai_settings) if unresolved else None
        ai_items, ai_messages = self._ai_parse_records(provider, unresolved)
        parsed = self._merge_recognized_items(local_items, ai_items)
        ai_used = bool(ai_items or ai_messages)
        resolved_keys = {(item["sheet_name"], item["row_number"]) for item in parsed}
        errors = [{"sheet": row["sheet"], "row": row["row"],
                   "message": "未能从该行识别出完整题干"}
                  for row in unresolved if (row["sheet"], row["row"]) not in resolved_keys]
        warnings = [{"sheet": item["sheet_name"], "row": item["row_number"], "message": message}
                    for item in parsed for message in item.get("recognition_notes", [])]
        warnings.extend({"sheet": "", "row": None, "message": message} for message in ai_messages)
        if unresolved and provider is None and ai_mode == "auto":
            warnings.append({"sheet": "", "row": None,
                             "message": "未配置可用智能服务；可填写教师自有 API 后重新导入以识别剩余内容"})

        unique: list[dict[str, Any]] = []
        seen_stems: set[str] = set()
        for item in sorted(parsed, key=lambda value: (value["sheet_name"], value["row_number"])):
            stem_key = re.sub(r"\s+", "", item["stem"]).lower()
            if stem_key in seen_stems:
                errors.append({"sheet": item["sheet_name"], "row": item["row_number"], "message": "题干重复"})
                continue
            seen_stems.add(stem_key)
            unique.append(item)
        if not unique:
            raise ValidationError("没有识别到可进入教师审核的题目；请检查文件内容或启用 API 辅助识别")
        return self._persist_flexible_import(
            actor, course_id, file_name, mime_type, data, digest, unique, errors, warnings,
            schemas, total_rows, ai_used, folder_id)

    def _persist_flexible_import(self, actor: dict[str, Any], course_id: str, file_name: str,
                                 mime_type: str, data: bytes, digest: str,
                                 items: list[dict[str, Any]], errors: list[dict[str, Any]],
                                 warnings: list[dict[str, Any]], schemas: list[dict[str, Any]],
                                 total_rows: int, ai_used: bool,
                                 folder_id: str | None) -> dict[str, Any]:
        import_id, document_id = f"qbi_{uuid.uuid4().hex}", f"doc_{uuid.uuid4().hex}"
        safe_name = _safe_name(file_name)
        destination_dir = self.campus.storage_dir / course_id / "question_bank"
        destination_dir.mkdir(parents=True, exist_ok=True)
        destination = destination_dir / f"{import_id}_{safe_name}"
        destination.write_bytes(data)
        parser_mode = "hybrid" if ai_used else "local"
        try:
            with self.db.connect() as conn:
                conn.execute(
                    """INSERT INTO course_documents(
                           document_id,course_id,uploader_id,original_name,stored_path,mime_type,
                           size_bytes,sha256,status
                       ) VALUES(?,?,?,?,?,?,?,?,'ready')""",
                    (document_id, course_id, actor["user_id"], safe_name, str(destination),
                     mime_type, len(data), digest),
                )
                conn.execute(
                    """INSERT INTO document_material_metadata(
                           document_id,material_type,suggested_material_type,classification_status,
                           tags_json,classification_reason,classified_by,classified_at
                       ) VALUES(?,'question_bank','question_bank','confirmed','[]',
                                '教师导入题库文件并完成内容识别',?,CURRENT_TIMESTAMP)""",
                    (document_id, actor["user_id"]),
                )
                conn.execute(
                    """INSERT INTO question_bank_imports(
                           import_id,course_id,document_id,original_name,stored_path,sha256,
                           total_rows,valid_rows,invalid_rows,errors_json,imported_by,
                           parser_mode,detected_schema_json,ai_used,warnings_json,folder_id
                       ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (import_id, course_id, document_id, safe_name, str(destination), digest,
                     total_rows, len(items), len(errors), _json(errors), actor["user_id"],
                     parser_mode, _json(schemas), int(ai_used), _json(warnings), folder_id),
                )
                for item in items:
                    conn.execute(
                        """INSERT INTO question_bank_items(
                               item_id,course_id,document_id,question_group_key,question_type,
                               stem_markdown,answer_markdown,explanation_markdown,
                               knowledge_points_json,source_pages_json,status,options_json,
                               correct_answer_json,difficulty,duration_seconds,source_kind,
                               import_row_number,import_id,recognition_confidence,
                               recognition_method,recognition_notes_json,folder_id
                           ) VALUES(?,?,?,?,?,?,?,?,?,?,'draft',?,?,?,?,?,?,?,?,?,?,?)""",
                        (f"qitem_{uuid.uuid4().hex}", course_id, document_id,
                         f"{import_id}:{item['sheet_name']}:{item['row_number']}",
                         item["question_type"], item["stem"], item["answer_markdown"],
                         item["explanation"], _json(item["knowledge_points"]), "[]",
                         _json(item["options"]), _json(item["answer"]), item["difficulty"],
                         item["duration_seconds"], "teacher_template", item["row_number"], import_id,
                         item["recognition_confidence"], item["recognition_method"],
                         _json(item.get("recognition_notes", [])), folder_id),
                    )
        except Exception:
            destination.unlink(missing_ok=True)
            raise
        result = self.db.fetch_one("SELECT * FROM question_bank_imports WHERE import_id=?", (import_id,)) or {}
        return {**self._format_import(result), "duplicate": False}

    def _parse_row(self, values: dict[str, Any], row_number: int) -> tuple[dict[str, Any] | None, str]:
        stem = _text(values.get("题干内容"))
        raw_answer = _text(values.get("答案"))
        question_type = self._question_type(
            _text(values.get("题目类型")), _text(values.get("答案类型")), raw_answer
        )
        if not stem:
            return None, "题干内容为空"
        if not raw_answer:
            return None, "标准答案为空，不能进入教师审核"

        options = [
            {"key": letter, "text": _text(values.get(f"选项{letter}"))}
            for letter in OPTION_LETTERS
            if _text(values.get(f"选项{letter}"))
        ]
        option_keys = {item["key"] for item in options}
        if question_type == "true_false":
            answer: str | list[str] = self._normalize_true_false(raw_answer)
            if not answer:
                return None, f"判断题答案“{raw_answer}”无法识别，应使用 T/F"
        elif question_type in {"single_choice", "multiple_choice"}:
            answers = self._choice_answers(raw_answer)
            if len(options) < 2:
                return None, "选择题至少需要两个非空选项"
            if not answers:
                return None, f"选择题答案“{raw_answer}”中没有 A-O 选项字母"
            missing = [value for value in answers if value not in option_keys]
            if missing:
                return None, f"答案引用了不存在的选项：{','.join(missing)}"
            if question_type == "single_choice" and len(answers) != 1:
                return None, "单选题必须且只能有一个正确选项"
            answer = answers if question_type == "multiple_choice" else answers[0]
        else:
            answer = raw_answer

        points = list(dict.fromkeys(filter(None, (
            _text(values.get("主题词一")),
            _text(values.get("主题词二")),
            _text(values.get("主题词三")),
            _text(values.get("知识点")),
        ))))
        duration = None
        raw_duration = values.get("答题时间（秒）")
        if raw_duration not in (None, ""):
            try:
                duration = max(1, min(86400, int(float(raw_duration))))
            except (TypeError, ValueError):
                return None, f"答题时间“{raw_duration}”不是有效秒数"
        return {
            "row_number": row_number,
            "question_type": question_type,
            "stem": stem,
            "answer": answer,
            "answer_markdown": ",".join(answer) if isinstance(answer, list) else answer,
            "explanation": _text(values.get("答案解析")),
            "options": options,
            "knowledge_points": points,
            "difficulty": _text(values.get("难度")),
            "duration_seconds": duration,
        }, ""

    def import_template(self, actor: dict[str, Any], course_id: str, file_name: str,
                        mime_type: str, data: bytes, *, ai_mode: str = "auto",
                        ai_settings: dict[str, str] | None = None,
                        folder_id: str | None = None) -> dict[str, Any]:
        return self.import_flexible(
            actor, course_id, file_name, mime_type, data,
            ai_mode=ai_mode, ai_settings=ai_settings, folder_id=folder_id,
        )
        self._teacher_course(actor, course_id)
        if Path(file_name).suffix.lower() != ".xlsx" or mime_type not in QUESTION_BANK_MIME_TYPES:
            raise ValidationError("正式题库仅支持 Excel XLSX 模板")
        if not data or len(data) > 20 * 1024 * 1024:
            raise ValidationError("题库文件不能为空且不能超过 20MB")
        digest = hashlib.sha256(data).hexdigest()
        existing = self.db.fetch_one(
            "SELECT * FROM question_bank_imports WHERE course_id=? AND sha256=?", (course_id, digest)
        )
        if existing:
            return {**self._format_import(existing), "duplicate": True}

        try:
            workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        except Exception as exc:
            raise ValidationError("Excel 文件损坏或不是有效的 XLSX 文件") from exc
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            header_values = next(rows, None)
            if not header_values:
                raise ValidationError("题库工作表为空")
            headers = [_text(value) for value in header_values]
            required = {"题目类型", "答案类型", "题干内容", "答案"}
            missing = sorted(required - set(headers))
            if missing:
                raise ValidationError(f"题库模板缺少必要列：{'、'.join(missing)}")

            parsed: list[dict[str, Any]] = []
            errors: list[dict[str, Any]] = []
            seen_stems: set[str] = set()
            total_rows = 0
            for row_number, row in enumerate(rows, 2):
                values = {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
                if not any(_text(value) for value in row):
                    continue
                total_rows += 1
                item, error = self._parse_row(values, row_number)
                if error:
                    errors.append({"row": row_number, "message": error})
                    continue
                assert item is not None
                stem_key = re.sub(r"\s+", "", item["stem"]).lower()
                if stem_key in seen_stems:
                    errors.append({"row": row_number, "message": "题干与本次导入中的其他题目重复"})
                    continue
                seen_stems.add(stem_key)
                parsed.append(item)
        finally:
            workbook.close()

        if not parsed:
            first_errors = "；".join(f"第{x['row']}行：{x['message']}" for x in errors[:3])
            raise ValidationError(f"没有可进入审核的有效题目。{first_errors}")

        import_id = f"qbi_{uuid.uuid4().hex}"
        document_id = f"doc_{uuid.uuid4().hex}"
        safe_name = _safe_name(file_name)
        destination_dir = self.campus.storage_dir / course_id / "question_bank"
        destination_dir.mkdir(parents=True, exist_ok=True)
        destination = destination_dir / f"{import_id}_{safe_name}"
        destination.write_bytes(data)

        try:
            with self.db.connect() as conn:
                conn.execute(
                    """INSERT INTO course_documents(
                           document_id,course_id,uploader_id,original_name,stored_path,mime_type,
                           size_bytes,sha256,status
                       ) VALUES(?,?,?,?,?,?,?,?,'ready')""",
                    (document_id, course_id, actor["user_id"], safe_name, str(destination),
                     mime_type, len(data), digest),
                )
                conn.execute(
                    """INSERT INTO document_material_metadata(
                           document_id,material_type,suggested_material_type,classification_status,
                           tags_json,classification_reason,classified_by,classified_at
                       ) VALUES(?,'question_bank','question_bank','confirmed','[]',
                                '教师通过题库模板导入',?,CURRENT_TIMESTAMP)""",
                    (document_id, actor["user_id"]),
                )
                conn.execute(
                    """INSERT INTO question_bank_imports(
                           import_id,course_id,document_id,original_name,stored_path,sha256,
                           total_rows,valid_rows,invalid_rows,errors_json,imported_by
                       ) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                    (import_id, course_id, document_id, safe_name, str(destination), digest,
                     total_rows, len(parsed), len(errors), _json(errors), actor["user_id"]),
                )
                for item in parsed:
                    item_id = f"qitem_{uuid.uuid4().hex}"
                    conn.execute(
                        """INSERT INTO question_bank_items(
                               item_id,course_id,document_id,question_group_key,question_type,
                               stem_markdown,answer_markdown,explanation_markdown,
                               knowledge_points_json,source_pages_json,status,options_json,
                               correct_answer_json,difficulty,duration_seconds,source_kind,
                               import_row_number,import_id
                           ) VALUES(?,?,?,?,?,?,?,?,?,?,'draft',?,?,?,?,?,?,?)""",
                        (
                            item_id, course_id, document_id, f"{import_id}:row:{item['row_number']}",
                            item["question_type"], item["stem"], item["answer_markdown"],
                            item["explanation"], _json(item["knowledge_points"]), "[]",
                            _json(item["options"]), _json(item["answer"]), item["difficulty"],
                            item["duration_seconds"], "teacher_template",
                            item["row_number"], import_id,
                        ),
                    )
        except Exception:
            destination.unlink(missing_ok=True)
            raise
        result = self.db.fetch_one("SELECT * FROM question_bank_imports WHERE import_id=?", (import_id,)) or {}
        return {**self._format_import(result), "duplicate": False}

    @staticmethod
    def _format_import(row: dict[str, Any]) -> dict[str, Any]:
        value = dict(row)
        value["errors"] = _loads(value.pop("errors_json", "[]"), [])
        value["warnings"] = _loads(value.pop("warnings_json", "[]"), [])
        value["detected_schema"] = _loads(value.pop("detected_schema_json", "[]"), [])
        value["ai_used"] = bool(value.get("ai_used"))
        value.pop("stored_path", None)
        value.pop("sha256", None)
        return value

    def list_imports(self, actor: dict[str, Any], course_id: str) -> list[dict[str, Any]]:
        self._teacher_course(actor, course_id)
        return [
            self._format_import(row)
            for row in self.db.fetch_all(
                "SELECT * FROM question_bank_imports WHERE course_id=? ORDER BY created_at DESC",
                (course_id,),
            )
        ]

    def latest_published(self, course_id: str,
                         folder_id: str | None = None) -> dict[str, Any] | None:
        if folder_id:
            return self.db.fetch_one(
                """SELECT * FROM question_bank_versions
                   WHERE course_id=? AND folder_id=? AND status='published'
                   ORDER BY version_number DESC LIMIT 1""",
                (course_id, folder_id),
            )
        return self.db.fetch_one(
            """SELECT * FROM question_bank_versions
               WHERE course_id=? AND folder_id IS NULL AND status='published'
               ORDER BY version_number DESC LIMIT 1""",
            (course_id,),
        )

    def student_publications(self, actor: dict[str, Any],
                             course_id: str) -> list[dict[str, Any]]:
        if actor.get("role") != "student":
            raise PermissionDenied("仅学生可以查看已发布练习")
        self.campus.require_access(course_id, str(actor["user_id"]), "student")
        return self.db.fetch_all(
            """SELECT f.folder_id,f.folder_name,f.folder_type,v.version_id,v.version_number,
                      COUNT(vi.item_id) item_count,v.published_at
               FROM question_bank_folders f
               JOIN question_bank_versions v ON v.folder_id=f.folder_id AND v.status='published'
               JOIN question_bank_version_items vi ON vi.version_id=v.version_id
               WHERE f.course_id=?
               GROUP BY v.version_id ORDER BY v.published_at DESC""",
            (course_id,),
        )

    def student_questions(self, actor: dict[str, Any], course_id: str, *,
                          limit: int = 30, offset: int = 0,
                          folder_id: str | None = None) -> dict[str, Any]:
        if actor.get("role") != "student":
            raise PermissionDenied("仅学生可以进入课程答题")
        course = self.campus.require_access(course_id, str(actor["user_id"]), "student")
        if course["course_type"] != "shared_course":
            raise ValidationError("教师发布题库仅用于共享课程")
        version = self.latest_published(course_id, folder_id)
        if not version:
            return {"version_id": None, "version_number": None, "total": 0, "items": []}
        limit = max(1, min(int(limit), 100))
        offset = max(0, int(offset))
        total = self.db.fetch_one(
            "SELECT COUNT(*) count FROM question_bank_version_items WHERE version_id=?",
            (version["version_id"],),
        )
        rows = self.db.fetch_all(
            """SELECT q.item_id,q.question_type,q.stem_markdown,q.options_json,
                      q.knowledge_points_json,q.difficulty,q.duration_seconds
               FROM question_bank_version_items vi
               JOIN question_bank_items q USING(item_id)
               WHERE vi.version_id=?
               ORDER BY q.import_row_number,q.created_at LIMIT ? OFFSET ?""",
            (version["version_id"], limit, offset),
        )
        items = []
        for row in rows:
            items.append({
                "item_id": row["item_id"],
                "type": row["question_type"],
                "question": row["stem_markdown"],
                "options": _loads(row["options_json"], []),
                "knowledge_points": _loads(row["knowledge_points_json"], []),
                "difficulty": row["difficulty"],
                "duration_seconds": row["duration_seconds"],
                "source_kind": "published_question_bank",
            })
        return {
            "version_id": version["version_id"],
            "version_number": version["version_number"],
            "folder_id": version.get("folder_id"),
            "total": int(total["count"] if total else 0),
            "items": items,
        }

    @classmethod
    def _is_correct(cls, question_type: str, response: Any, answer: Any) -> bool:
        if question_type == "true_false":
            normalized_response = cls._normalize_true_false(_text(response))
            normalized_answer = cls._normalize_true_false(_text(answer))
            return bool(normalized_response) and normalized_response == normalized_answer
        if question_type == "multiple_choice":
            submitted = response if isinstance(response, list) else cls._choice_answers(_text(response))
            return sorted({_text(value).upper() for value in submitted if _text(value)}) == sorted(answer)
        if question_type == "single_choice":
            selected = _text(response).upper()
            return selected == _text(answer).upper()
        normalize = lambda value: re.sub(r"[\s，,。；;]+", "", _text(value)).lower()
        return bool(normalize(response)) and normalize(response) == normalize(answer)

    def submit(self, actor: dict[str, Any], course_id: str, version_id: str,
               responses: list[dict[str, Any]]) -> dict[str, Any]:
        if actor.get("role") != "student":
            raise PermissionDenied("仅学生可以提交题库答案")
        self.campus.require_access(course_id, str(actor["user_id"]), "student")
        current = self.db.fetch_one(
            """SELECT * FROM question_bank_versions
               WHERE version_id=? AND course_id=? AND status='published'""",
            (version_id, course_id),
        )
        if not current:
            raise ValidationError("题库已更新，请刷新后重新作答")
        if not responses or len(responses) > 100:
            raise ValidationError("每次需提交 1 至 100 道题")
        item_ids = list(dict.fromkeys(_text(item.get("item_id")) for item in responses))
        if any(not value for value in item_ids) or len(item_ids) != len(responses):
            raise ValidationError("提交中包含空题号或重复题目")
        placeholders = ",".join("?" for _ in item_ids)
        rows = self.db.fetch_all(
            f"""SELECT q.* FROM question_bank_version_items vi
                JOIN question_bank_items q USING(item_id)
                WHERE vi.version_id=? AND q.item_id IN ({placeholders})""",
            (version_id, *item_ids),
        )
        by_id = {row["item_id"]: row for row in rows}
        if len(by_id) != len(item_ids):
            raise ValidationError("提交中包含不属于当前发布版本的题目")

        submission_id = f"qsub_{uuid.uuid4().hex}"
        results = []
        with self.db.connect() as conn:
            for submitted in responses:
                item = by_id[_text(submitted["item_id"])]
                response = submitted.get("response", "")
                answer = _loads(item["correct_answer_json"], item["answer_markdown"])
                correct = self._is_correct(item["question_type"], response, answer)
                conn.execute(
                    """INSERT INTO question_bank_attempts(
                           attempt_id,submission_id,course_id,version_id,item_id,student_id,
                           response_json,is_correct
                       ) VALUES(?,?,?,?,?,?,?,?)""",
                    (f"qat_{uuid.uuid4().hex}", submission_id, course_id, version_id,
                     item["item_id"], actor["user_id"], _json(response), int(correct)),
                )
                results.append({
                    "item_id": item["item_id"],
                    "correct": correct,
                    "response": response,
                    "correct_answer": answer,
                    "explanation": item["explanation_markdown"],
                })
        correct_count = sum(bool(item["correct"]) for item in results)
        return {
            "submission_id": submission_id,
            "correct": correct_count,
            "total": len(results),
            "accuracy": round(correct_count * 100 / len(results), 1),
            "results": results,
        }

    def statistics(self, actor: dict[str, Any], course_id: str,
                   class_id: str | None = None,
                   folder_id: str | None = None) -> dict[str, Any]:
        self._teacher_course(actor, course_id)
        member_condition = ""
        params: list[Any] = [course_id]
        class_info = None
        if class_id:
            class_info = self.db.fetch_one(
                "SELECT * FROM classes WHERE class_id=? AND course_id=? AND teacher_id=?",
                (class_id, course_id, actor["user_id"]),
            )
            if not class_info:
                raise PermissionDenied("无权查看该教学班题库统计")
            member_condition = "AND a.student_id IN (SELECT student_id FROM class_memberships WHERE class_id=? AND status='active')"
            params.append(class_id)

        version = self.latest_published(course_id, folder_id)
        if not version:
            return {
                "version": None, "summary": {"students": 0, "answered": 0, "accuracy": 0},
                "ranking": [], "students": [],
            }
        latest_cte = f"""
            WITH ranked AS (
                SELECT a.*,ROW_NUMBER() OVER (
                    PARTITION BY a.student_id,a.item_id ORDER BY a.submitted_at DESC,a.attempt_id DESC
                ) position
                FROM question_bank_attempts a
                WHERE a.course_id=? {member_condition}
            ), latest AS (SELECT * FROM ranked WHERE position=1)
        """
        ranking_rows = self.db.fetch_all(
            latest_cte + """
            SELECT q.item_id,q.stem_markdown,q.question_type,
                   COUNT(l.attempt_id) attempts,
                   COALESCE(SUM(l.is_correct),0) correct_count,
                   COALESCE(SUM(CASE WHEN l.is_correct=0 THEN 1 ELSE 0 END),0) wrong_count
            FROM question_bank_version_items vi
            JOIN question_bank_items q USING(item_id)
            LEFT JOIN latest l ON l.item_id=q.item_id
            WHERE vi.version_id=?
            GROUP BY q.item_id,q.stem_markdown,q.question_type
            ORDER BY CASE WHEN COUNT(l.attempt_id)=0 THEN -1
                          ELSE 1.0*SUM(CASE WHEN l.is_correct=0 THEN 1 ELSE 0 END)/COUNT(l.attempt_id)
                     END DESC,COUNT(l.attempt_id) DESC,q.import_row_number
            """,
            (*params, version["version_id"]),
        )
        ranking = []
        for index, row in enumerate(ranking_rows, 1):
            attempts = int(row["attempts"])
            wrong = int(row["wrong_count"])
            ranking.append({
                **row,
                "rank": index if attempts else None,
                "error_rate": round(wrong * 100 / attempts, 1) if attempts else 0,
                "accuracy": round(int(row["correct_count"]) * 100 / attempts, 1) if attempts else 0,
            })

        if class_id:
            members = self.db.fetch_all(
                """SELECT m.student_id,u.student_number,u.display_name,u.username
                   FROM class_memberships m LEFT JOIN users u ON u.user_id=m.student_id
                   WHERE m.class_id=? AND m.status='active' ORDER BY u.student_number,u.display_name""",
                (class_id,),
            )
        else:
            members = self.db.fetch_all(
                """SELECT DISTINCT e.student_id,u.student_number,u.display_name,u.username
                   FROM course_enrollments e LEFT JOIN users u ON u.user_id=e.student_id
                   WHERE e.course_id=? ORDER BY u.student_number,u.display_name""",
                (course_id,),
            )
        latest_rows = self.db.fetch_all(
            latest_cte + """
            SELECT l.student_id,l.item_id,l.is_correct,l.response_json,l.submitted_at,
                   q.stem_markdown,q.answer_markdown
            FROM latest l JOIN question_bank_items q USING(item_id)
            JOIN question_bank_version_items vi ON vi.item_id=q.item_id
            WHERE vi.version_id=?
            """,
            (*params, version["version_id"]),
        )
        by_student: dict[str, list[dict[str, Any]]] = {}
        for row in latest_rows:
            by_student.setdefault(row["student_id"], []).append(row)
        students = []
        for member in members:
            attempts = by_student.get(member["student_id"], [])
            wrong = [{
                "item_id": row["item_id"],
                "question": row["stem_markdown"],
                "response": _loads(row["response_json"], ""),
                "correct_answer": row["answer_markdown"],
                "submitted_at": row["submitted_at"],
            } for row in attempts if not row["is_correct"]]
            correct = sum(bool(row["is_correct"]) for row in attempts)
            students.append({
                **member,
                "answered": len(attempts),
                "correct": correct,
                "wrong_count": len(wrong),
                "accuracy": round(correct * 100 / len(attempts), 1) if attempts else 0,
                "wrong_questions": wrong,
            })
        answered_students = sum(item["answered"] > 0 for item in students)
        total_attempts = sum(item["answered"] for item in students)
        total_correct = sum(item["correct"] for item in students)
        return {
            "version": version,
            "class": class_info,
            "summary": {
                "students": len(students),
                "answered": answered_students,
                "attempts": total_attempts,
                "accuracy": round(total_correct * 100 / total_attempts, 1) if total_attempts else 0,
            },
            "ranking": ranking,
            "students": students,
        }
